"""File I/O, outbox, upload, workspace CRUD, and file search handlers."""

from __future__ import annotations

import asyncio
import contextlib
import datetime as _dt
import errno
import hashlib
import json
import logging
import mimetypes
import ntpath
import os
import re
import stat as _stat_mod
import subprocess
import sys
import threading
import time
import urllib.parse
import uuid
import zipfile
from pathlib import Path
from typing import BinaryIO, NamedTuple

from aiohttp import web
from aiohttp.client_exceptions import ClientConnectionResetError
from aiohttp.multipart import BodyPartReader

from kiro_crew import file_delivery_consent, pinned_fs, platform_compat
from kiro_crew.atomic_write import (
    atomic_write,
    open_access_control_source,
    pinned_parent_replace_supported,
)
from kiro_crew.config import loader as config_loader
from kiro_crew.config.loader import KiroCrewConfig, WorkspaceConfig, config_dir, data_home
from kiro_crew.dashboard import part_stream, upload_destination
from kiro_crew.dashboard.chat_utils import dashboard_slot_key
from kiro_crew.dashboard.file_index import _SKIP_DIRS as _WALK_SKIP_DIRS
from kiro_crew.dashboard.handlers._shared import _probe_persisted_session, read_bounded_json
from kiro_crew.dashboard.origin import is_direct_local_request
from kiro_crew.dashboard.state import DashboardState, append_and_surface
from kiro_crew.doc_parser import extract_text
from kiro_crew.hooks import FileTooLargeError, safe_read_file_bytes, safe_read_prefix
from kiro_crew.messaging.display_safety import redact_for_display
from kiro_crew.messaging.outbound_files import OutboundFile
from kiro_crew.messaging.raster import SNIFF_BYTES, sniff_raster_mime
from kiro_crew.platform import redact_via_context as redact
from kiro_crew.sandbox import popen_limited, sandboxed_spawn_argv
from kiro_crew.security import (
    BINARY_MIME_ALLOWLIST,
    is_sensitive_path,
    redact_credentials,
    redact_exfiltration_urls,
)
from kiro_crew.slack.handler import is_tracked_channel
from kiro_crew.validation import (
    FILE_READ_SCHEMA,
    ValidationError,
    validate_tool_args,
)
from kiro_crew.zip_vet import ZipInventoryRejected, vet_zip_inventory_bytes

# Register OOXML office MIME types explicitly. The system mimetypes
# database on AL2/AL2023 build hosts does NOT include .docx, .xlsx, or
# .pptx by default, so mimetypes.guess_type() returns (None, None) for
# those. Registering at module import time keeps api_file_download's
# Content-Type header correct for the most common Word/Excel/PowerPoint
# downloads.
mimetypes.add_type(
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document", ".docx",
)
mimetypes.add_type(
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", ".xlsx",
)
mimetypes.add_type(
    "application/vnd.openxmlformats-officedocument.presentationml.presentation", ".pptx",
)

_INLINE_DISPOSITION_PREFIXES = frozenset({"audio/", "video/", "image/", "application/pdf"})


logger = logging.getLogger(__name__)


def _sel():
    """Late-binding _sel() for test monkeypatch compatibility."""
    import kiro_crew.dashboard.handlers as _pkg  # noqa: F811
    return _pkg.sel()


def _audit_file_send(
    *,
    leg: str,
    outcome: str,
    error: str | None = None,
    downstream: str | None = None,
    resources: str | None = None,
) -> None:
    """The one audit shape both ``file_send`` delivery legs write.

    Every record the Slack and channel endpoints emit is the same tool
    invocation under a different ``tool_kind`` (the leg), so the shape lives
    here rather than being spelled out at each of the dozen decision sites it
    used to be copied to -- one drifted field was previously a one-line edit
    away. Optional fields are OMITTED when unset, exactly as the shipped call
    sites omitted them: skips carry no ``downstream_service``, refusals and
    deliveries do.
    """
    extra: dict[str, str] = {}
    if error is not None:
        extra["error"] = error
    if downstream is not None:
        extra["downstream_service"] = downstream
    if resources is not None:
        extra["resources"] = resources
    _sel().log_tool_invocation(
        session_key="api",
        source="api",
        tool_name="file_send",
        tool_kind=leg,
        outcome=outcome,
        **extra,
    )


def _body_err_code(body_err: web.Response) -> str:
    """SEL error label for a refused body read.

    Derived from the guard response's machine-readable ``code`` so the audit
    record distinguishes a parse failure from an oversized body (413
    ``payload_too_large``) instead of filing every refusal as a JSON error.
    """
    try:
        parsed = json.loads(body_err.text or "")
    except ValueError:
        return "invalid_json_body"
    code = parsed.get("code") if isinstance(parsed, dict) else None
    return str(code) if code else "invalid_json_body"


async def api_reveal_path(request: web.Request) -> web.Response:
    """POST /api/reveal — reveal a file/folder in Finder or open with default app."""
    # Default cap: the body is a path and an action flag (issue #5587 sweep).
    body, body_err = await read_bounded_json(request)
    if body_err is not None:
        return body_err
    assert body is not None  # read_bounded_json returns (dict, None) on success
    path = body.get("path", "")
    action = body.get("action", "reveal")  # "reveal" or "open"
    if not path or ".." in Path(path).parts:
        return web.json_response({"error": "invalid path"}, status=400)
    if is_sensitive_path(path):
        _sel().log_tool_invocation(
            session_key="api", source="api", tool_name="reveal_path",
            outcome="denied", error="sensitive_path",
            resources=path, metadata={"action": action})
        return web.json_response({"error": "access denied"}, status=403)
    # Gate: only spawn native openers from direct-local requests. Remote/tunneled
    # callers get the copy-to-clipboard fallback — spawning Finder on a machine
    # the user is not looking at is surprising and useless.
    if not is_direct_local_request(request):
        _sel().log_tool_invocation(
            session_key="api", source="api", tool_name="reveal_path",
            outcome="denied", error="remote_request",
            resources=path, metadata={"action": action})
        # Degrade to a clipboard copy: `copy` is the path to write. The remote
        # cause is recorded in the SEL audit above (error="remote_request"); the
        # response body carries no path, host, or exception detail beyond `copy`.
        return web.json_response({"ok": True, "copy": path})
    # Every ALLOWED outcome leaves through the single audited return below —
    # including the clipboard answer, which is a granted decision whose host
    # simply had no file manager. An early return here would drop that decision
    # from the SEL log, so the branches record what happened instead of exiting.
    #
    # Both spawns live in platform_compat, which owns the safety properties:
    # absolute trusted launchers rather than bare argv names, a folder rather
    # than the file on the platforms where handing a file to the file manager
    # would launch it, and Windows refused outright for the launch-by-association
    # verb. They answer False both for a host with no launcher and for one that
    # refuses to start, and either way this degrades to the clipboard rather than
    # failing a click in the file viewer.
    if action == "open":
        if not os.path.isfile(path):
            return web.json_response({"error": "not a regular file"}, status=400)
        copied = not platform_compat.open_with_default_app(path)
    else:
        copied = not platform_compat.reveal_in_file_manager(path)
    _sel().log_tool_invocation(
        session_key="api", source="api", tool_name="reveal_path",
        outcome="success", resources=path, metadata={"action": action})
    # A local grant whose host had no working file manager degrades to the
    # clipboard; `copy` is the path to write.
    if copied:
        return web.json_response({"ok": True, "copy": path})
    return web.json_response({"ok": True})


async def api_outbox_notify(request: web.Request) -> web.Response:
    """POST /api/outbox/notify — agent sent a file, notify the user."""
    state: DashboardState = request.app["state"]
    # Default cap: the body names an outbox file (path, filename, short
    # description, size) — the file bytes themselves never travel in it.
    body, body_err = await read_bounded_json(request)
    if body_err is not None:
        _sel().log_tool_invocation(
            session_key="api",
            source="api",
            tool_name="file_send",
            tool_kind="notify",
            outcome="denied",
            error=_body_err_code(body_err),
        )
        return body_err
    assert body is not None  # read_bounded_json returns (dict, None) on success

    raw_path = body.get("path", "")
    raw_filename = body.get("filename", "")
    raw_desc = body.get("description", "")
    # Reject files whose names/paths contain sensitive patterns
    if redact(raw_filename) != raw_filename or redact(raw_path) != raw_path:

        _sel().log_tool_invocation(
            session_key="api",
            source="api",
            tool_name="file_send",
            tool_kind="notify",
            outcome="denied",
            error="sensitive_filename_rejected",
        )
        return web.json_response(
            {"error": "filename or path contains sensitive content"}, status=400
        )
    file_data = {
        "filename": raw_filename,
        "path": raw_path,
        "description": redact(raw_desc),
        "size": body.get("size", 0),
        "content_type": mimetypes.guess_type(raw_filename)[0] or "application/octet-stream",
    }
    # Validate file is readable + UTF-8 before creating a persistent card
    from pathlib import Path  # noqa: F811

    from kiro_crew.config.loader import outbox_dir  # noqa: F811
    from kiro_crew.hooks import FileTooLargeError, safe_read_file_bytes  # noqa: F811

    resolved = Path(file_data["path"]).resolve()
    if not resolved.is_relative_to(outbox_dir().resolve()):

        _sel().log_tool_invocation(
            session_key="api",
            source="api",
            tool_name="file_send",
            tool_kind="notify",
            outcome="denied",
            error="path_outside_outbox",
        )
        return web.json_response({"error": "path must be inside outbox"}, status=403)
    try:
        raw = safe_read_file_bytes(str(resolved))
    except FileTooLargeError as e:

        _sel().log_tool_invocation(
            session_key="api",
            source="api",
            tool_name="file_send",
            tool_kind="notify",
            outcome="denied",
            error=f"file_too_large: {e}",
        )
        return web.json_response({"error": str(e)}, status=413)
    if raw is None:

        _sel().log_tool_invocation(
            session_key="api",
            source="api",
            tool_name="file_send",
            tool_kind="notify",
            outcome="denied",
            error="file_not_found_or_access_denied",
        )
        return web.json_response({"error": "File not found or access denied"}, status=404)
    # Text files: check for sensitive content. Binary files: skip content scan
    # and validate MIME against the shared BINARY_MIME_ALLOWLIST.
    try:
        text = raw.decode("utf-8")
        # The owner's grant covers this leg: the card renders in the owner's own
        # authenticated dashboard. No audit event here -- the delivery decision is
        # already recorded by the tool leg, and the byte handover is recorded by
        # the download route; a third entry for rendering a card would only bury
        # the two that answer a real question.
        if redact(text) != text and not file_delivery_consent.is_granted(
            file_delivery_consent.CLASS_OWNER_DASHBOARD
        ):
            _sel().log_tool_invocation(
                session_key="api",
                source="api",
                tool_name="file_send",
                tool_kind="notify",
                outcome="denied",
                error="sensitive_content_detected",
            )
            return web.json_response({"error": "file content contains sensitive data"}, status=400)
    except UnicodeDecodeError:
        # Binary file — only allow known-safe media types
        guessed_type = mimetypes.guess_type(raw_filename)[0] or ""
        if guessed_type not in BINARY_MIME_ALLOWLIST:
            _sel().log_tool_invocation(
                session_key="api",
                source="api",
                tool_name="file_send",
                tool_kind="notify",
                outcome="denied",
                error=f"binary_mime_not_allowed: {guessed_type}",
            )
            return web.json_response(
                {"error": f"Binary file type not allowed: {guessed_type or 'unknown'}"}, status=400
            )
    # Inject into the caller's chat slot so the card persists in the correct session
    if state._slots:
        # Prefer the caller's own slot via X-Session-Key header
        session_key = request.headers.get("X-Session-Key", "").strip()
        active = None
        if session_key.startswith("cron:"):
            # A cron slot is named cron-<id>, which is not the session key folded.
            active = state.get_slot(f"cron-{session_key.removeprefix('cron:')}")
        else:
            # A channel-born conversation keeps its channel key (slack:<ts>)
            # while its tab is open, so the slot name comes from the surface
            # lookup — stripping a "dashboard:" prefix would miss it and drop the
            # card into whichever tab happened to be active last.
            slot_key = dashboard_slot_key(session_key)
            if slot_key:
                active = state.get_slot(slot_key)
        # An explicitly header-targeted slot receives the file even when empty
        header_targeted = active is not None
        # Fallback: most recently active slot
        if not active:
            active = max(
                state._slots.values(),
                key=lambda s: s.messages[-1]["ts"] if s.messages else "",
            )
        if active and (active.messages or header_targeted):
            # Route through the context-aware redact() so a loaded companion's
            # extra credential regexes scrub the broadcast file JSON too — the
            # same overlay-aware pass the filename/path/description gates use.
            redacted_file_json = redact(json.dumps(file_data))
            # append_and_surface = the same conditional-broadcast pattern this
            # site pioneered, now also stamping ``ts`` + ``meta.mid`` on the
            # reader-suppressed frame so a client seeing the row through two
            # doors recognises it instead of rendering a duplicate card.
            append_and_surface(state, active, "file", redacted_file_json)

    _sel().log_tool_invocation(
        session_key="api",
        source="api",
        tool_name="file_send",
        tool_kind="notify",
        outcome="completed",
        resources=f"filename={file_data['filename']}",
    )
    return web.json_response({"ok": True})


async def api_outbox_download(request: web.Request) -> web.StreamResponse:
    """GET /api/outbox/{filename} — download a file from the outbox."""
    from kiro_crew.config.loader import outbox_dir  # noqa: F811
    from kiro_crew.hooks import FileTooLargeError, safe_read_file_bytes  # noqa: F811

    filename = request.match_info["filename"]
    path = (outbox_dir() / filename).resolve()
    if not path.is_relative_to(outbox_dir().resolve()):
        _sel().log_tool_invocation(
            session_key="api",
            source="api",
            tool_name="file_send",
            tool_kind="download",
            outcome="denied",
            error=f"path_traversal: {filename}",
        )
        return web.json_response({"error": "forbidden"}, status=403)
    try:
        raw = safe_read_file_bytes(str(path))
    except FileTooLargeError as e:
        _sel().log_tool_invocation(
            session_key="api",
            source="api",
            tool_name="file_send",
            tool_kind="download",
            outcome="denied",
            error=f"file_too_large: {e}",
        )
        return web.json_response({"error": str(e)}, status=413)
    if raw is None:
        _sel().log_tool_invocation(
            session_key="api",
            source="api",
            tool_name="file_send",
            tool_kind="download",
            outcome="denied",
            error=f"safe_read_file_bytes rejected: {filename}",
        )
        return web.json_response({"error": "forbidden"}, status=403)
    # For text files, scan for sensitive content; binary files served as-is
    # against the shared BINARY_MIME_ALLOWLIST (deny-by-default).
    is_text = True
    try:
        text = raw.decode("utf-8")
    except UnicodeDecodeError:
        is_text = False
    if is_text:
        redacted = redact(text)
        if redacted != text:
            # This is where the flagged bytes actually leave for the owner's
            # browser, so a grant is honoured here AND the handover is audited --
            # the refusal it replaces was self-evident in the 400, whereas a
            # successful consented download would otherwise leave no trace.
            #
            # TWO conjuncts, and the second is not redundant. This route is absent
            # from every ``token_auth`` bypass list, which establishes that it needs
            # AUTHENTICATION -- not that it needs OWNER IDENTITY. A Slack
            # allow-listed non-owner running ``!dashboard`` authenticates with
            # ``app == ""`` and ``sub != owner_id``, so ordinary token auth admits
            # them while ``is_owner_dashboard_request`` does not. Without the owner
            # conjunct the grant would convert a clean 400-for-everyone into raw
            # bytes for every authenticated caller -- widening the audience as a
            # side effect of a control meant to narrow it, and contradicting the
            # "owner's own authenticated browser" audience this class is scoped to.
            from kiro_crew.dashboard.handlers.source_providers import (  # lazy: import cycle
                is_owner_dashboard_request,
            )

            if not (
                file_delivery_consent.is_granted(file_delivery_consent.CLASS_OWNER_DASHBOARD)
                and is_owner_dashboard_request(request)
            ):
                _sel().log_tool_invocation(
                    session_key="api",
                    source="api",
                    tool_name="file_send",
                    tool_kind="download",
                    outcome="denied",
                    error="content_redacted",
                )
                return web.json_response(
                    {"error": "file content was redacted; download aborted"}, status=400
                )
            _sel().log_tool_invocation(
                session_key="api",
                source="api",
                tool_name="file_send",
                tool_kind="download",
                outcome="completed",
                error="sensitive_content_delivered_with_consent",
            )
            file_delivery_consent.audit_decision(
                file_delivery_consent.CLASS_OWNER_DASHBOARD,
                outcome="delivered",
                detail=f"download: {path.name}",
            )
    safe_name = urllib.parse.quote(path.name, safe="")
    content_type, _ = mimetypes.guess_type(path.name)
    if not content_type:
        content_type = "application/octet-stream"
    # Binary files must be in the allowlist
    if not is_text and content_type not in BINARY_MIME_ALLOWLIST:
        _sel().log_tool_invocation(
            session_key="api",
            source="api",
            tool_name="file_send",
            tool_kind="download",
            outcome="denied",
            error=f"binary_mime_not_allowed: {content_type}",
        )
        return web.json_response(
            {"error": f"Binary file type not allowed: {content_type}"}, status=403
        )
    # Inline disposition for media types the browser can render
    disposition = "inline" if any(content_type.startswith(t) for t in _INLINE_DISPOSITION_PREFIXES) else "attachment"
    # SVG can contain scripts — never serve inline on the dashboard origin
    if content_type == "image/svg+xml":
        disposition = "attachment"
    # Text files always attachment — prevents content injection via crafted filenames
    if is_text:
        disposition = "attachment"
    _sel().log_tool_invocation(
        session_key="api",
        source="api",
        tool_name="file_send",
        tool_kind="download",
        outcome="completed",
        resources=f"filename={filename}",
    )
    return web.Response(
        body=raw,
        headers={
            "Content-Disposition": f"{disposition}; filename*=UTF-8''{safe_name}",
            "Content-Type": content_type,
            "X-Content-Type-Options": "nosniff",
        },
    )


async def api_outbox_list(request: web.Request) -> web.Response:
    """GET /api/outbox — list files in the outbox."""
    from kiro_crew.config.loader import outbox_dir  # noqa: F811

    entries = []
    odir = outbox_dir()
    if not odir.is_dir():
        return web.json_response({"files": []})
    for f in odir.iterdir():
        try:
            st = f.stat()
        except FileNotFoundError:
            continue
        if f.is_file() and redact(f.name) == f.name:
            entries.append({"filename": f.name, "size": st.st_size, "modified": st.st_mtime})
    entries.sort(key=lambda x: float(x["modified"]), reverse=True)  # type: ignore[arg-type,return-value]

    _sel().log_tool_invocation(
        session_key="api",
        source="api",
        tool_name="file_send",
        tool_kind="list",
        outcome="completed",
        resources=f"count={len(entries)}",
    )
    return web.json_response({"files": entries[:50]})


def _gate_upload_file(
    file_path: str, filename: str, *, tool_kind: str
) -> tuple[web.Response | None, Path | None, bytes | None]:
    """The shared admission gate for shipping a local file to a channel.

    One site computes the judgment for every channel-upload endpoint —
    containment (outbox or workspace root), the descriptor-safe read, the
    binary MIME allowlist, and the content credential scans — so the Slack
    and channel legs cannot drift apart gate by gate. Returns
    ``(error_response, None, None)`` on refusal, ``(None, resolved, bytes)``
    when the file may ship. *tool_kind* keys the SEL records so each caller
    keeps its own audit lane.

    Blocking by design (a full read of up to ``MAX_FILE_BYTES`` plus content
    regex scans): async handlers MUST run it off the event loop via
    ``asyncio.to_thread`` — SEL appends are internally locked, so the audit
    calls are thread-safe. The loader is called through its module so tests
    (and config reloads) resolve at call time, not import time.
    """

    def _audit_denial(error: str, *, outcome: str = "denied") -> None:
        _sel().log_tool_invocation(
            session_key="api",
            source="api",
            tool_name="file_send",
            tool_kind=tool_kind,
            outcome=outcome,
            downstream_service=tool_kind,
            error=error,
        )

    if not file_path or not filename:
        _audit_denial("missing_required_fields")
        return (
            web.json_response(
                {"error": "file_path, filename required", "code": "missing_required_fields"},
                status=400,
            ),
            None,
            None,
        )
    # The name is DELIVERED (Slack upload title, Telegram document name,
    # Discord message text fallback), so a credential embedded in it leaves
    # with the file. Checked in the shared gate so no leg can drift from the
    # others, and before path resolution so a sensitive name never even
    # selects a file. Mirrors the MCP-side file_send refusal.
    if redact(filename) != filename:
        _audit_denial("sensitive_filename_rejected")
        return (
            web.json_response(
                {
                    "error": "filename contains sensitive content",
                    "code": "sensitive_filename",
                },
                status=400,
            ),
            None,
            None,
        )
    resolved = Path(file_path).resolve()
    allowed_outbox = config_loader.outbox_dir().resolve()
    allowed_workspace = config_loader.workspace_root().resolve()
    if not (resolved.is_relative_to(allowed_outbox) or resolved.is_relative_to(allowed_workspace)):
        _audit_denial(f"path_not_allowed: {file_path}")
        return (
            web.json_response(
                {
                    "error": "file_path must be under the outbox directory or the workspace root",
                    "code": "path_not_allowed",
                },
                status=403,
            ),
            None,
            None,
        )
    try:
        raw = safe_read_file_bytes(str(resolved))
    except FileTooLargeError as e:
        _audit_denial(f"file_too_large: {e}")
        return (
            web.json_response({"error": str(e), "code": "file_too_large"}, status=413),
            None,
            None,
        )
    if raw is None:
        _audit_denial(f"safe_read_file_bytes rejected: {file_path}")
        return (
            web.json_response(
                {
                    "error": f"File not found or access denied: {file_path}",
                    "code": "file_not_found",
                },
                status=404,
            ),
            None,
            None,
        )
    try:
        text = raw.decode("utf-8")
    except UnicodeDecodeError:
        # Binary file — only allow known-safe media types
        guessed_type = mimetypes.guess_type(filename)[0] or ""
        if guessed_type not in BINARY_MIME_ALLOWLIST:
            _audit_denial(f"binary_mime_not_allowed: {guessed_type}")
            return (
                web.json_response(
                    {
                        "error": f"Binary file type not allowed: {guessed_type or 'unknown'}",
                        "code": "binary_mime_not_allowed",
                    },
                    status=400,
                ),
                None,
                None,
            )
        text = None  # signal: skip text redaction path
        # Scan binary content for embedded credentials (e.g. base64-encoded keys in PDFs)
        binary_text = raw.decode("latin-1")
        if redact(binary_text) != binary_text:
            _audit_denial("binary_credential_detected")
            return (
                web.json_response(
                    {
                        "error": "binary file contains embedded credentials",
                        "code": "binary_credential_detected",
                    },
                    status=400,
                ),
                None,
                None,
            )
    if text is not None:
        try:
            redacted = redact(text)
            if redacted != text:
                _audit_denial("content_redacted")
                return (
                    web.json_response(
                        {
                            "error": "file content was redacted; upload aborted",
                            "code": "content_redacted",
                        },
                        status=400,
                    ),
                    None,
                    None,
                )
        except Exception as redact_err:
            _audit_denial(f"redaction_failed: {redact_err}", outcome="error")
            return (
                web.json_response(
                    {"error": f"Redaction failed: {redact_err}", "code": "redaction_failed"},
                    status=500,
                ),
                None,
                None,
            )
    return None, resolved, raw


async def api_slack_upload_file(request: web.Request) -> web.Response:
    """POST /api/slack/upload-file — upload a file to Slack (internal, called by file_send).

    Destination and authorization come from the shared oracle
    (:func:`kiro_crew.dashboard.upload_destination.resolve_slack`), which holds
    this leg's ladder — the ``channels``-scope governance vet, the
    restricted-session ceiling, then a request-named channel, a
    session-map-linked thread, or the owner-DM fallback with its tracked-channel
    authorization — next to the non-Slack leg's, so the two cannot drift apart
    rung by rung (issue #6060). What stays here is what only this leg can
    answer: the Slack client, its upload verb, and the response shapes.

    The client-presence check stays AHEAD of the body parse, where it shipped: a
    gateway with no Slack client answers ``skipped: no_slack`` even for a
    malformed body.
    """
    state: DashboardState = request.app["state"]
    slack = state.slack_client
    if not slack:
        _audit_file_send(leg="slack", outcome="skipped", error="no_slack_client")
        return web.json_response({"ok": True, "skipped": "no_slack"})
    # Default cap: the body carries a file path, a filename, and Slack routing
    # ids — the file bytes are read from disk, never from this body.
    body, body_err = await read_bounded_json(request)
    if body_err is not None:
        _audit_file_send(leg="slack", outcome="denied", error=_body_err_code(body_err))
        return body_err
    assert body is not None  # read_bounded_json returns (dict, None) on success
    file_path_raw = body.get("file_path", "")
    filename = body.get("filename", "")
    # Off-loop: the gate reads up to MAX_FILE_BYTES and regex-scans the content
    # (no-blocking-call-on-event-loop).
    error_resp, resolved, raw = await asyncio.to_thread(
        _gate_upload_file, file_path_raw, filename, tool_kind="slack"
    )
    if error_resp is not None:
        return error_resp
    assert resolved is not None and raw is not None  # narrowed by the gate
    # ``is_tracked_channel`` and the persisted-transcript probe are handed to the
    # oracle rather than imported there: one binding site, and the module stays
    # free of both the Slack handler's config dependency and the ``dashboard``
    # package ``messaging.upload_gate`` may not import.
    destination = await upload_destination.resolve_slack(
        state,
        slack,
        session_key=request.headers.get("X-Session-Key", "").strip(),
        requested_channel=body.get("channel", ""),
        thread_ts=body.get("thread_ts"),
        tracked_probe=is_tracked_channel,
        persisted_probe=_probe_persisted_session,
    )
    if isinstance(destination, upload_destination.Refusal):
        _audit_file_send(
            leg="slack",
            outcome="denied",
            error=destination.audit_error,
            downstream=destination.downstream,
        )
        # One branch per literal status, body inline. `status=<expression>` and a
        # body hoisted into a variable are both invisible to the error-code
        # contract scanner, which counts either as its own bucket
        # (test_error_code_contract) -- so the refusal says WHICH answer it is
        # and each answer is spelled out here.
        if destination.status == 400:
            return web.json_response(
                {"error": destination.error, "code": destination.code}, status=400
            )
        return web.json_response(
            {"error": destination.error, "code": destination.code}, status=403
        )
    if isinstance(destination, upload_destination.Skip):
        _audit_file_send(leg="slack", outcome="skipped", error=destination.reason)
        return web.json_response({"ok": True, "skipped": destination.reason})
    try:
        # The filename was already cleared by the shared admission gate above —
        # same predicate, same value, strictly earlier in this function — so the
        # leg no longer re-checks it. #6044 made that gate the one site for the
        # rule; a second copy here could only drift from it.
        await slack.upload_file(
            destination.channel,
            destination.thread_ts,
            str(resolved),
            filename,
            filename,
        )
        _audit_file_send(
            leg="slack",
            outcome="completed",
            downstream="slack",
            resources=f"channel={destination.channel} file={file_path_raw}",
        )
        return web.json_response({"ok": True})
    except Exception as e:
        # A Slack SDK / network exception can carry file paths, host and URL
        # fragments, or credentials embedded in a URL. Sanitize before it
        # reaches the client or the audit record (see api_slack_pins).
        safe_error, _ = redact_credentials(str(e))
        safe_error, _ = redact_exfiltration_urls(safe_error)
        _audit_file_send(leg="slack", outcome="error", downstream="slack", error=safe_error)
        return web.json_response({"error": safe_error}, status=500)


async def api_channel_upload_file(request: web.Request) -> web.Response:
    """POST /api/channel/upload-file — deliver a file to the caller's own
    conversation on a non-Slack channel (internal, called by file_send).

    Destination and authorization come from the shared oracle
    (:func:`kiro_crew.dashboard.upload_destination.resolve_channel`), which for
    this leg is the SAME send ladder the cross-surface reply mirror uses
    (``_resolve_mirror_target``): channel-scope governance, transport
    registration, proactive-send capability, and ``may_send_to`` recipient
    re-authorization, all fail-closed and SEL-audited in one place — plus the
    restricted-session ceiling the renderers' extraction path enforces, on the
    same shared predicate. The destination comes exclusively from the caller's
    session map entry — a request cannot name an arbitrary conversation, which is
    what keeps this endpoint from being a broadcast primitive. The oracle also
    resolves the delivery verb, since which channels have one is part of "can
    this file land here": Telegram and Discord today, each via its own
    purpose-built name-preserving ``send_document``; every other channel is a
    skip until its transport grows that verb. The Slack counterpart above
    resolves through the same module, one rung table away (issue #6060).

    "Cannot deliver here" is a SKIP (``delivered: false``), not an error: most
    sessions mirror nowhere, and the caller falls back to the dashboard card
    and the Slack leg exactly as before this endpoint existed.
    """
    state: DashboardState = request.app["state"]
    # Default cap: same shape as the Slack leg — a path, a filename, and a
    # short description; the file bytes are read from disk by the gate.
    body, body_err = await read_bounded_json(request)
    if body_err is not None:
        _audit_file_send(leg="channel", outcome="denied", error=_body_err_code(body_err))
        return body_err
    assert body is not None  # read_bounded_json returns (dict, None) on success

    def _skip(reason: str) -> web.Response:
        _audit_file_send(leg="channel", outcome="skipped", error=reason)
        return web.json_response({"ok": True, "delivered": False, "skipped": reason})

    destination = await upload_destination.resolve_channel(
        state,
        request.headers.get("X-Session-Key", "").strip(),
        persisted_probe=_probe_persisted_session,
    )
    if isinstance(destination, upload_destination.Skip):
        return _skip(destination.reason)
    link, deliver = destination.link, destination.deliver
    # Off-loop: the gate reads up to MAX_FILE_BYTES and regex-scans the content
    # (no-blocking-call-on-event-loop).
    error_resp, resolved, raw = await asyncio.to_thread(
        _gate_upload_file,
        body.get("file_path", ""),
        body.get("filename", ""),
        tool_kind="channel",
    )
    if error_resp is not None:
        return error_resp
    assert resolved is not None and raw is not None  # narrowed by the gate
    filename = body.get("filename", "")
    # Display-form redaction, not just literal: redact() scans bytes, and the
    # channel's renderer strips markup at display time — ``AKIA**…**`` passes
    # a literal scan and displays as an intact key. Same boundary rule every
    # renderer sink applies (``redact_for_display``) before text reaches a
    # transport.
    description, _ = redact_for_display(body.get("description", "") or "", redact)
    outbound = OutboundFile(
        path=str(resolved),
        data=raw,
        alt=description,
        mime=mimetypes.guess_type(filename)[0] or "application/octet-stream",
    )
    try:
        mid = await deliver(
            link.channel_id,
            outbound,
            caption=description,
            thread_id=link.thread_id,
        )
    except Exception as e:
        # A transport / network exception can carry file paths, host and URL
        # fragments, or credentials embedded in a URL. Sanitize before it
        # reaches the client or the audit record (see api_slack_upload_file).
        safe_error, _ = redact_credentials(str(e))
        safe_error, _ = redact_exfiltration_urls(safe_error)
        _audit_file_send(
            leg="channel",
            outcome="error",
            downstream=link.channel_type,
            error=safe_error,
        )
        return web.json_response({"error": safe_error}, status=502)
    if not mid:
        # The transport reported failure without raising (the clients return
        # an empty id on an API-level refusal).
        _audit_file_send(
            leg="channel",
            outcome="error",
            downstream=link.channel_type,
            error="delivery_reported_no_message_id",
        )
        return web.json_response({"error": "channel delivery failed"}, status=502)
    _audit_file_send(
        leg="channel",
        outcome="completed",
        downstream=link.channel_type,
        resources=f"channel_type={link.channel_type} file={body.get('file_path', '')}",
    )
    return web.json_response(
        {"ok": True, "delivered": True, "channel_type": link.channel_type}
    )


async def api_upload(request: web.Request) -> web.Response:
    """POST /api/upload — open native file picker and return selected paths."""
    if sys.platform != "darwin":
        return web.json_response({"error": "File picker is only available on macOS"}, status=400)

    proc = await asyncio.create_subprocess_exec(
        "osascript",
        "-e",
        "set f to choose file with multiple selections allowed\n"
        'set out to ""\n'
        "repeat with p in f\n"
        "  set out to out & POSIX path of p & linefeed\n"
        "end repeat\n"
        "return out",
        stdout=asyncio.subprocess.PIPE,
        stderr=asyncio.subprocess.DEVNULL,
    )
    try:
        stdout, _ = await asyncio.wait_for(proc.communicate(), timeout=120)
    except asyncio.TimeoutError:
        try:
            proc.kill()
        except ProcessLookupError:
            pass
        await proc.communicate()
        return web.json_response({"error": "Finder dialog timed out"}, status=504)
    paths = [ln for ln in stdout.decode("utf-8", errors="replace").strip().splitlines() if ln]

    if not paths:
        return web.json_response({"paths": []})
    return web.json_response({"paths": paths})


# Resolved per call, never captured at import: an import-time binding freezes
# the data home and defeats pod isolation, the lazy legacy-home migration and
# test isolation. The name below is an opt-in override (None = live home) so
# existing monkeypatch call sites keep working. See config.md "Data Home";
# dashboard/handlers/usage.py is the reference implementation.
_SCREENSHOT_DIR: Path | None = None

_UPLOAD_DIR: Path | None = None


def _screenshot_dir() -> Path:
    """Screenshots directory, resolved against the live data home."""
    return _SCREENSHOT_DIR if _SCREENSHOT_DIR is not None else data_home() / "screenshots"


def _upload_dir() -> Path:
    """Uploads directory, resolved against the live data home."""
    return _UPLOAD_DIR if _UPLOAD_DIR is not None else data_home() / "uploads"


_MAX_UPLOAD_BYTES = 50 * 1024 * 1024  # 50 MB per file
#: Video gets its own, larger ceiling: a 30-second retina screen recording is
#: routinely 60-150 MB, so the 50 MB document cap would reject the dominant
#: case and make the feature read as broken. Safe to raise only because video
#: parts STREAM to disk (:func:`_stream_video_part`) instead of accumulating in
#: memory the way every other accepted type does.
_MAX_VIDEO_UPLOAD_BYTES = 512 * 1024 * 1024  # 512 MB per video
_MAX_UPLOAD_FILES = 20  # max files per request

# Fallback-walk budgets. ``_WALK_MAX_SCAN_*`` bounds entries scored PER KIND
# (anti-starvation); ``_WALK_MAX_DIRS_VISITED`` bounds directories entered and is
# what guarantees termination -- see ``_walk_file_search``. Not a multiple of the
# per-kind budget: in a narrow-deep tree directory names grow at the same rate as
# directories visited, so a derived ceiling is unreachable exactly when it is
# needed. Module-level so tests can shrink them.
_WALK_MAX_SCAN_SCOPED = 50_000
_WALK_MAX_SCAN_UNSCOPED = 5_000
_WALK_MAX_DIRS_VISITED = 20_000

# Hard ceiling on the caller-supplied ``limit`` of /api/file-search. The walk
# collects ``max_results * 10`` candidates per kind, so the limit multiplies real
# filesystem work; a fixed server-side ceiling keeps a hostile ``?limit=`` from
# turning the endpoint into a filesystem-walk amplifier. Mirrored client-side as
# SEARCH_RESULT_LIMIT_MAX in FolderPanel.tsx.
_SEARCH_LIMIT_CEILING = 60
_ALLOWED_IMAGE_EXT = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp", ".svg"}
_ALLOWED_TEXT_EXT = {
    ".txt",
    ".md",
    ".json",
    # Excalidraw scene JSON — the composer's sketch pad attaches one per
    # sketch, and the dashboard has a dedicated read-only renderer for it
    # (FileRenderers routes on this exact extension). Content-wise it is
    # ordinary JSON text.
    ".excalidraw",
    ".har",
    ".yaml",
    ".yml",
    ".xml",
    ".csv",
    ".log",
    ".py",
    ".js",
    ".ts",
    ".tsx",
    ".jsx",
    ".html",
    ".css",
    ".sh",
    ".bash",
    ".rb",
    ".go",
    ".rs",
    ".java",
    ".c",
    ".cpp",
    ".h",
    ".hpp",
}
_ALLOWED_DOC_EXT = {
    ".pdf",
    ".doc",
    ".docx",
    ".xls",
    ".xlsx",
    ".ppt",
    ".pptx",
    ".odt",
    ".ods",
    ".odp",
    ".rtf",
    ".zip",
    ".tar",
    ".gz",
}
#: Video containers accepted at the upload boundary. Deliberately narrower than
#: ``FileRenderers``' VIDEO_EXTS: every entry here must be verifiable by
#: :func:`_sniff_media_type` AND playable by ``<video>``, so an accepted upload
#: is always one the chat can actually show. ``.mkv`` is excluded — it shares
#: WebM's EBML signature but browser playback is unreliable, and accepting a
#: file that then refuses to play is worse than refusing it at the door.
_ALLOWED_VIDEO_EXT = {".mp4", ".m4v", ".mov", ".webm"}
#: Media containers a browser will often play but the upload boundary does not
#: accept. Rejecting them with the bare "Unsupported file type" reads as "video
#: is not supported at all", when the actual remedy is a re-encode -- so the
#: refusal for one of these names the containers that do work. VIDEO containers
#: only: naming the video set to an audio upload (``.m4a``) would tell its
#: sender to re-encode audio into a video container, which is worse than the
#: bare refusal.
_VIDEO_HINT_EXT = frozenset(
    {".mkv", ".ogv", ".avi", ".mpg", ".mpeg", ".wmv", ".flv", ".3gp"}
)
#: Media type :func:`_sniff_media_type` must report for the claimed video
#: extension. The MP4 family (mp4/m4v/mov) all carry a ``ftyp`` box at offset 4
#: and sniff as ``video/mp4``; QuickTime's brand differs but the box does not.
#:
#: This gate proves the bytes are the claimed FAMILY, not the exact container:
#: ``.webm`` and ``.mkv`` share the EBML magic, so an ``.mkv`` renamed to
#: ``.webm`` passes here even though the ``.mkv`` extension is refused.
#: Distinguishing them needs the EBML DocType, which is not worth parsing for
#: this boundary -- the gate's job is to keep NON-media bytes off disk (CWE-434),
#: and the extension set is what carries the narrower "accepted means playable"
#: promise.
_VIDEO_EXT_MIME: dict[str, str] = {
    ".mp4": "video/mp4",
    ".m4v": "video/mp4",
    ".mov": "video/mp4",
    ".webm": "video/webm",
}


def _write_file_restricted(path: Path, data: bytes) -> None:
    """Write file with owner-only permissions (0o600)."""
    fd = os.open(
        str(path),
        os.O_WRONLY | os.O_CREAT | os.O_TRUNC | getattr(os, "O_BINARY", 0),
        0o600,
    )
    try:
        os.write(fd, data)
    finally:
        os.close(fd)


def _open_rb_nofollow(path: str) -> int:
    """Open *path* read-only in binary, refusing symlinks, on every platform.

    POSIX gets the atomic form: ``O_NOFOLLOW`` makes the kernel itself fail
    the open with ``ELOOP`` when the final component is a symlink, so there is
    no check-then-open race. Windows has no ``O_NOFOLLOW`` (referencing it
    raises AttributeError, turning every read into an HTTP 500), so there the
    guard is a pre-open ``lstat``: reject symlinks and any reparse point
    (junctions included) with the same ``ELOOP`` errno the POSIX branch
    produces, keeping callers' error handling identical. The window between
    lstat and open is acceptable defence-in-depth there -- path containment
    was already enforced by the caller's validation, and creating a symlink
    on Windows requires elevated or developer-mode privileges. ``O_BINARY``
    keeps the CRT from text-mode translating file bytes on Windows; it is 0
    elsewhere.
    """
    nofollow = getattr(os, "O_NOFOLLOW", 0)
    if not nofollow:
        st = os.lstat(path)
        if _stat_mod.S_ISLNK(st.st_mode) or getattr(st, "st_reparse_tag", 0) != 0:
            raise OSError(errno.ELOOP, "symlinks not allowed", path)
    return os.open(path, os.O_RDONLY | nofollow | getattr(os, "O_BINARY", 0))


# Magic-byte signatures for content-type validation at the upload boundary
# (CWE-434). The extension is attacker-controlled, so binary types are verified
# against their file signature BEFORE the bytes are written. Raster types are
# verified by the shared sniffer (:mod:`kiro_crew.messaging.raster`), so all
# consumers agree on what counts as each image type (including WebP's form tag
# at offset 8, which a bare ``RIFF`` prefix would not check). Text formats (and
# SVG, which is XML) have no reliable magic and remain gated by the extension
# allowlist only.
_ZIP_CONTAINER_EXTS = {".docx", ".xlsx", ".pptx", ".odt", ".ods", ".odp", ".zip"}
#: Raster extensions and the mime :func:`sniff_raster_mime` must report for
#: the claimed extension to be accepted.
_RASTER_EXT_MIME: dict[str, str] = {
    ".png": "image/png",
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".gif": "image/gif",
    ".bmp": "image/bmp",
    ".webp": "image/webp",
}
#: Non-raster binary types that still carry a reliable leading signature.
_MAGIC_PREFIXES: dict[str, tuple[bytes, ...]] = {
    ".pdf": (b"%PDF-",),
    ".gz": (b"\x1f\x8b",),
}
#: Read-path extras the shared raster table does not cover (served by
#: ``api_file_raw`` but never accepted at the upload boundary).
_READ_PATH_EXTRA_MAGIC: tuple[tuple[bytes, str], ...] = (
    (b"II\x2a\x00", "image/tiff"),
    (b"MM\x00\x2a", "image/tiff"),
    (b"\x00\x00\x01\x00", "image/x-icon"),
)


def _content_matches_ext(ext: str, data: bytes) -> bool:
    """Best-effort magic-byte check that ``data`` matches the claimed ``ext``.

    Returns False only when the signature is KNOWN and does not match, so an
    attacker can't store arbitrary bytes (e.g. an HTML/script payload) under an
    allowed binary extension (CWE-434). Unknown / text extensions (and ``.svg``)
    return True — there is no reliable signature — and stay gated by the
    extension allowlist alone.
    """
    if ext in _ZIP_CONTAINER_EXTS:
        # OOXML / ODF / zip all begin with a local-file-header, empty-archive,
        # or spanned-archive PK signature.
        return data[:4] in (b"PK\x03\x04", b"PK\x05\x06", b"PK\x07\x08")
    expected_media = _VIDEO_EXT_MIME.get(ext)
    if expected_media is not None:
        # Reuses the read path's container sniffer so the upload boundary and
        # /api/file-stream agree on what each signature means. ``data`` may be
        # just the leading chunk here — every signature involved lives in the
        # first 12 bytes, so a header is sufficient and a whole-file read is
        # never needed.
        return _sniff_media_type(data[:SNIFF_BYTES]) == expected_media
    expected = _RASTER_EXT_MIME.get(ext)
    if expected is not None:
        return sniff_raster_mime(data[:SNIFF_BYTES]) == expected
    prefixes = _MAGIC_PREFIXES.get(ext)
    if prefixes is None:
        return True  # text / svg / unknown — nothing to enforce
    return any(data.startswith(p) for p in prefixes)


async def _stream_video_part(
    part: BodyPartReader,
    dest: Path,
) -> tuple[int, tuple[str, str, str] | None]:
    """Stream a video *part* to *dest*, gating on its container signature.

    Returns ``(bytes_written, None)`` on success, or ``(bytes_written,
    (audit_reason, error_code, user_message))`` on refusal. The code is a
    machine-readable id the caller maps to a CONSTANT HTTP status: returning a
    status from here would make the response's `status=` an expression at the
    call site, which the error-code contract rejects because it defeats static
    analysis of what the endpoint can return.

    All the file handling lives in :func:`~kiro_crew.dashboard.part_stream.
    stream_part_to_file`, which owns the temp through a synchronous context
    manager. This function is now only the translation between that helper's
    exceptions and this endpoint's audit reasons and error codes -- deliberately,
    because the hand-rolled version of the streaming here collected SEVEN
    blocking review findings in seven rounds, three of them introduced while
    fixing the previous one. That module's docstring carries the ledger and the
    invariant; the short version is that a cancellable coroutine cannot own a
    file safely, so it no longer does.
    """
    ext = dest.suffix.lower()
    try:
        total = await part_stream.stream_part_to_file(
            part,
            dest,
            max_bytes=_MAX_VIDEO_UPLOAD_BYTES,
            accepts=lambda head: _content_matches_ext(ext, head),
        )
    except part_stream.PartTooLarge as too_large:
        cap_mb = _MAX_VIDEO_UPLOAD_BYTES // 1024 // 1024
        return too_large.total, (
            f"too_large:{too_large.total}",
            "video_too_large",
            f"Video too large (max {cap_mb}MB)",
        )
    except part_stream.PartContentMismatch:
        accepted = ", ".join(sorted(_ALLOWED_VIDEO_EXT))
        return 0, (
            f"content_signature_mismatch:{ext}",
            "video_content_mismatch",
            # Names the remedy for the same reason the unsupported-container
            # refusal does: "does not match its type" tells the user their file
            # is wrong without telling them what to do about it, and the fix
            # (re-export) is not guessable from the sentence.
            f"This file is not really a {ext} — re-export it as one of: {accepted}",
        )
    return total, None


async def api_upload_file(request: web.Request) -> web.Response:
    """POST /api/upload/file — cross-platform multipart file upload.

    Accepts multipart form data with one or more 'file' fields.
    Saves files to the data home's uploads/ and returns server-side paths
    that ACP's _send_prompt() can detect for image inlining.
    """

    upload_dir = _upload_dir()
    upload_dir.mkdir(parents=True, exist_ok=True)
    reader = await request.multipart()
    paths: list[str] = []
    allowed = _ALLOWED_IMAGE_EXT | _ALLOWED_TEXT_EXT | _ALLOWED_DOC_EXT | _ALLOWED_VIDEO_EXT
    caller = request.get("user", "dashboard")

    async def _cleanup(*also: Path) -> None:
        """Remove this request's files, plus *also*, off the serving loop.

        The ONE cleanup entry point for this handler, and a coroutine so it
        cannot be called the blocking way by accident. Every refusal and error
        path in a 20-file request may unlink up to 20 paths (a video among them
        up to 512 MB), and `Path.unlink` is a synchronous syscall: on a slow or
        network filesystem doing that inline stalls chat and heartbeat for the
        whole gateway. It also absorbs the destination itself via *also*, so the
        sites that previously paired a bare ``dest.unlink()`` with a cleanup call
        have one call and cannot drift back to unlinking on the loop.
        """
        targets = [*paths, *(str(p) for p in also)]

        def _rm() -> None:
            for p in targets:
                Path(p).unlink(missing_ok=True)

        await asyncio.to_thread(_rm)

    try:
        while True:
            part = await reader.next()
            if part is None:
                break
            if not isinstance(part, BodyPartReader):
                continue
            if part.name != "file":
                continue
            if len(paths) >= _MAX_UPLOAD_FILES:
                await _cleanup()
                _sel().log_api_access(
                    caller=caller,
                    operation="upload.file",
                    outcome="rejected",
                    source="dashboard",
                    resources=f"reason:too_many_files:{_MAX_UPLOAD_FILES}",
                )
                return web.json_response(
                    {"error": f"Too many files (max {_MAX_UPLOAD_FILES})"},
                    status=400,
                )
            fname = part.filename or "upload"
            # Sanitize: strip path components to prevent traversal
            safe_name = re.sub(r"[^\w.\-]", "_", Path(fname).name)
            ext = Path(safe_name).suffix.lower()
            if ext not in allowed:
                await _cleanup()
                _sel().log_api_access(
                    caller=caller,
                    operation="upload.file",
                    outcome="rejected",
                    source="dashboard",
                    resources=f"file:{fname} reason:unsupported_type:{ext}",
                )
                detail = f"Unsupported file type: {ext}"
                if ext in _VIDEO_HINT_EXT:
                    # Name the way out. A browser plays several containers this
                    # boundary refuses, so the bare refusal reads as "no video
                    # support" when the remedy is a re-encode.
                    accepted = ", ".join(sorted(_ALLOWED_VIDEO_EXT))
                    detail = f"{detail} — accepted video containers: {accepted}"
                return web.json_response(
                    {"error": detail, "code": "unsupported_file_type"},
                    status=400,
                )
            # UUID prefix guarantees uniqueness even within a single request.
            # Resolved BEFORE any byte is read because the video branch streams
            # straight to this destination rather than buffering the part first.
            dest = upload_dir / f"{uuid.uuid4().hex}_{safe_name}"
            if not dest.resolve().is_relative_to(upload_dir.resolve()):
                await _cleanup()
                _sel().log_api_access(
                    caller=caller,
                    operation="upload.file",
                    outcome="rejected",
                    source="dashboard",
                    resources=f"file:{fname} reason:path_traversal",
                )
                return web.json_response({"error": "Invalid filename"}, status=400)
            if ext in _ALLOWED_VIDEO_EXT:
                # Video takes the streaming route for two reasons: a screen
                # recording is far too large to buffer, and its CONTENT is not
                # something the model can read anyway (ACP has no video content
                # block). So the bytes land on disk, the PATH reaches the agent
                # as an [attached_file N] token, and the chat renders a <video>
                # off /api/file-stream. An agent that needs frames runs ffmpeg
                # on the path.
                try:
                    written, refusal = await _stream_video_part(part, dest)
                except (Exception, asyncio.CancelledError):
                    # CancelledError derives from BaseException, not Exception, so
                    # a bare `except Exception` lets a gateway shutdown mid-stream
                    # past every cleanup: the partial video AND the siblings this
                    # request already wrote stay in uploads/, and the partial is
                    # indistinguishable from a complete file to everything
                    # downstream. Cleanup here rather than relying on the outer
                    # handler, which has the same blind spot.
                    await _cleanup(dest)
                    raise
                if refusal is not None:
                    await _cleanup(dest)
                    reason, code, message = refusal
                    _sel().log_api_access(
                        caller=caller,
                        operation="upload.file",
                        outcome="rejected",
                        source="dashboard",
                        resources=f"file:{fname} reason:{reason}",
                    )
                    # Branched rather than parameterised: each response states a
                    # CONSTANT status and its own `code`, which is what keeps the
                    # endpoint's possible outcomes statically readable (and is
                    # what the error-code contract checks for).
                    if code == "video_too_large":
                        return web.json_response(
                            {"error": message, "code": "video_too_large"},
                            status=413,
                        )
                    return web.json_response(
                        {"error": message, "code": "video_content_mismatch"},
                        status=400,
                    )
                logger.info(
                    "upload.file video: name=%s ext=%s size=%d",
                    safe_name,
                    ext,
                    written,
                )
                paths.append(str(dest))
                continue
            # Read with size limit
            data = bytearray()
            while True:
                chunk = await part.read_chunk(8192)
                if not chunk:
                    break
                data.extend(chunk)
                if len(data) > _MAX_UPLOAD_BYTES:
                    await _cleanup()
                    _sel().log_api_access(
                        caller=caller,
                        operation="upload.file",
                        outcome="rejected",
                        source="dashboard",
                        resources=f"file:{fname} reason:too_large:{len(data)}",
                    )
                    return web.json_response(
                        {"error": f"File too large (max {_MAX_UPLOAD_BYTES // 1024 // 1024}MB)"},
                        status=413,
                    )
            # Content-signature gate (CWE-434): verify magic bytes match the
            # claimed extension BEFORE writing, so an allowed extension can't
            # smuggle arbitrary/binary content (e.g. a .png that is really HTML).
            if not _content_matches_ext(ext, bytes(data)):
                await _cleanup()
                _sel().log_api_access(
                    caller=caller,
                    operation="upload.file",
                    outcome="rejected",
                    source="dashboard",
                    resources=f"file:{fname} reason:content_signature_mismatch:{ext}",
                )
                return web.json_response(
                    {"error": f"File content does not match its type: {ext}"},
                    status=400,
                )
            try:
                await asyncio.to_thread(_write_file_restricted, dest, bytes(data))
            except Exception:
                await _cleanup(dest)
                raise
            # Diagnostic logging for binary uploads. Compares the bytes
            # we received in memory against the bytes that landed on
            # disk after _write_file_restricted, so a future report of
            # "uploaded .docx is corrupted" can be pinned to the
            # upload pipeline vs post-upload tampering. Logged for
            # extensions that are binary archives (docx/xlsx/pptx/odt/
            # zip/pdf etc.) where any byte mismatch breaks the file;
            # text uploads aren't worth the I/O.
            if ext in _ALLOWED_DOC_EXT or ext in _ALLOWED_IMAGE_EXT:
                try:
                    sent_sha = hashlib.sha256(bytes(data)).hexdigest()
                    on_disk = dest.read_bytes()
                    disk_sha = hashlib.sha256(on_disk).hexdigest()
                    head_hex = on_disk[:4].hex() if on_disk else ""
                    is_zip_ext = ext in {".docx", ".xlsx", ".pptx", ".odt", ".ods", ".odp", ".zip"}
                    is_zip = zipfile.is_zipfile(str(dest)) if is_zip_ext else None
                    logger.info(
                        "upload.file diagnostic: name=%s ext=%s sent_size=%d disk_size=%d "
                        "sent_sha256=%s disk_sha256=%s match=%s magic=%s is_zipfile=%s",
                        safe_name,
                        ext,
                        len(data),
                        len(on_disk),
                        sent_sha,
                        disk_sha,
                        sent_sha == disk_sha,
                        head_hex,
                        is_zip,
                    )
                except Exception:
                    # Diagnostic failure must never break the upload.
                    logger.exception("upload.file diagnostic failed for %s", safe_name)
            paths.append(str(dest))
    except (Exception, asyncio.CancelledError):
        # Same blind spot as the video branch above: a cancelled request (gateway
        # shutdown, client disconnect) raises CancelledError, which is NOT an
        # Exception, so without naming it every file this request already wrote
        # is orphaned in uploads/ with nothing left to reference or remove it.
        await _cleanup()
        _sel().log_api_access(
            caller=caller,
            operation="upload.file",
            outcome="error",
            source="dashboard",
            resources=f"files_written:{len(paths)}",
        )
        raise
    if not paths:
        _sel().log_api_access(
            caller=caller,
            operation="upload.file",
            outcome="rejected",
            source="dashboard",
            resources="reason:no_files",
        )
        return web.json_response({"error": "No files uploaded"}, status=400)
    _sel().log_api_access(
        caller=caller,
        operation="upload.file",
        outcome="success",
        source="dashboard",
        resources=f"files:{len(paths)}",
    )
    return web.json_response({"paths": paths})


async def api_screenshot(request: web.Request) -> web.Response:
    """POST /api/screenshot — capture screen region and return file path.

    macOS only — uses built-in screencapture. Linux cloud desktops
    (AL2, headless) don't have a display server so this is unavailable.
    """
    if sys.platform != "darwin":
        return web.json_response({"error": "Screenshot is only available on macOS"}, status=400)

    screenshot_dir = _screenshot_dir()
    screenshot_dir.mkdir(parents=True, exist_ok=True)
    ts = int(time.time())
    dest = screenshot_dir / f"screenshot_{ts}.png"

    proc = await asyncio.create_subprocess_exec(
        "screencapture",
        "-i",
        str(dest),
        stdout=asyncio.subprocess.DEVNULL,
        stderr=asyncio.subprocess.DEVNULL,
    )
    try:
        await asyncio.wait_for(proc.wait(), timeout=120)
    except asyncio.TimeoutError:
        try:
            proc.kill()
        except ProcessLookupError:
            pass
        await proc.wait()
        return web.json_response({"error": "screenshot timed out"}, status=504)
    if not dest.exists():
        return web.json_response({"path": ""})  # user cancelled
    return web.json_response({"path": str(dest)})


# ── Workspace API ──
async def api_workspaces(request: web.Request) -> web.Response:
    """GET /api/workspaces — list configured workspaces."""
    cfg = KiroCrewConfig.load()
    default_ws = cfg.default_workspace
    result = []
    for name, ws in cfg.workspaces.items():
        result.append({"name": name, "path": ws.dir, "is_default": name == default_ws})
    if not result:
        result.append({"name": "default", "path": "workspace", "is_default": True})
    return web.json_response({"workspaces": result, "default": default_ws})


async def api_workspaces_create(request: web.Request) -> web.Response:
    """POST /api/workspaces — create a new workspace."""
    import shutil  # noqa: F811

    from kiro_crew.dashboard.handlers._shared import require_owner_dashboard_request
    from kiro_crew.validation import WORKSPACE_NAME_RE  # noqa: F811

    # Ahead of the body read: a workspace entry carries a caller-supplied
    # directory, so the traversal and sensitive-path guards below are defending
    # against input that only the owner may supply in the first place.
    owner_denied = await require_owner_dashboard_request(request, "workspace.create")
    if owner_denied is not None:
        return owner_denied

    # Default cap: the body is a workspace name plus optional dir/copy_from.
    body, body_err = await read_bounded_json(request)
    if body_err is not None:
        return body_err
    assert body is not None  # read_bounded_json returns (dict, None) on success
    name = body.get("name", "").strip()
    if not name:
        return web.json_response({"error": "Workspace name is required"}, status=400)
    if not WORKSPACE_NAME_RE.match(name):
        return web.json_response(
            {"error": "Invalid workspace name (use alphanumeric, hyphens, underscores)"},
            status=400,
        )
    cfg = KiroCrewConfig.load()
    if name in cfg.workspaces:
        return web.json_response({"error": f"Workspace '{name}' already exists"}, status=409)
    copy_from = body.get("copy_from", "").strip()
    if copy_from:
        if copy_from not in cfg.workspaces:
            return web.json_response(
                {"error": f"Source workspace '{copy_from}' not found"}, status=404
            )
        # New workspace gets its own directory, named after the workspace
        ws_dir = body.get("dir", f"workspace-{name}")
        # Check for directory collision with existing workspaces
        existing_dirs = {ws.dir for ws in cfg.workspaces.values()}
        if ws_dir in existing_dirs:
            return web.json_response(
                {"error": f"Directory '{ws_dir}' is already used by another workspace"},
                status=409,
            )
        # Recursively copy source workspace data to the new directory
        src_path = data_home() / cfg.workspaces[copy_from].dir
        dst_path = data_home() / ws_dir
        # Guard against path traversal
        if not dst_path.resolve().is_relative_to(data_home().resolve()):
            _sel().log_api_access(
                caller=request.get("user", "dashboard"),
                operation="workspace.create",
                outcome="denied",
                source="dashboard",
                resources=name,
            )
            return web.json_response({"error": "Invalid directory path"}, status=400)
        if not src_path.resolve().is_relative_to(data_home().resolve()):
            _sel().log_api_access(
                caller=request.get("user", "dashboard"),
                operation="workspace.create",
                outcome="denied",
                source="dashboard",
                resources=name,
            )
            return web.json_response({"error": "Invalid source directory path"}, status=400)
        # Reject config root itself to avoid copying .env / config.json
        cfg_root = data_home().resolve()
        if src_path.resolve() == cfg_root or dst_path.resolve() == cfg_root:
            _sel().log_api_access(
                caller=request.get("user", "dashboard"),
                operation="workspace.create",
                outcome="denied",
                source="dashboard",
                resources=name,
            )
            return web.json_response(
                {"error": "Cannot use config root as workspace directory"}, status=400
            )
        if src_path.is_dir():
            # Use the module-level is_sensitive_path alias to filter entries
            # instead of hardcoded names -- one binding for one guard.

            def _ignore_sensitive(directory: str, entries: list[str]) -> set[str]:
                from pathlib import Path as _Path  # noqa: F811

                skip: set[str] = set()
                for entry in entries:
                    full = str(_Path(directory, entry).resolve())
                    if is_sensitive_path(full):
                        skip.add(entry)
                return skip

            await asyncio.to_thread(
                shutil.copytree,
                src_path,
                dst_path,
                dirs_exist_ok=True,
                symlinks=True,
                ignore=_ignore_sensitive,
            )
    else:
        ws_dir = body.get("dir", f"workspace-{name}")
    # Guard against path traversal for relative paths; absolute paths are allowed
    _abs = Path(ws_dir).expanduser().is_absolute()
    # Path constructed for validation only (never opened/read/written); the
    # is_relative_to + is_sensitive_path guards below reject traversals before
    # the value is stored in config. CodeQL's taint tracker does not model the
    # containment guard as a barrier.
    final_path = (  # lgtm[py/path-injection]
        Path(ws_dir).expanduser().resolve() if _abs else data_home() / ws_dir
    )

    # Check for directory collision with existing workspaces (resolve both sides)
    def _resolve_ws_dir(d: str) -> Path:
        p = Path(d).expanduser()
        return p.resolve() if p.is_absolute() else (data_home() / d).resolve()

    existing_resolved = {_resolve_ws_dir(ws.dir) for ws in cfg.workspaces.values()}
    if _resolve_ws_dir(ws_dir) in existing_resolved:
        return web.json_response(
            {"error": f"Directory '{ws_dir}' is already used by another workspace"},
            status=409,
        )
    if is_sensitive_path(str(final_path.resolve())):
        _sel().log_api_access(
            caller=request.get("user", "dashboard"),
            operation="workspace.create",
            outcome="denied",
            source="dashboard",
            resources=name,
        )
        return web.json_response({"error": "Invalid directory path"}, status=400)
    if not _abs and not final_path.resolve().is_relative_to(data_home().resolve()):
        _sel().log_api_access(
            caller=request.get("user", "dashboard"),
            operation="workspace.create",
            outcome="denied",
            source="dashboard",
            resources=name,
        )
        return web.json_response({"error": "Invalid directory path"}, status=400)
    if final_path.resolve() == data_home().resolve():
        _sel().log_api_access(
            caller=request.get("user", "dashboard"),
            operation="workspace.create",
            outcome="denied",
            source="dashboard",
            resources=name,
        )
        return web.json_response(
            {"error": "Cannot use config root as workspace directory"}, status=400
        )
    cfg.workspaces[name] = WorkspaceConfig(dir=ws_dir)
    cfg.save()
    _sel().log_api_access(
        caller=request.get("user", "dashboard"),
        operation="workspace.create",
        outcome="success",
        source="dashboard",
        resources=name,
    )
    return web.json_response({"ok": True, "name": name})


async def api_workspaces_update(request: web.Request) -> web.Response:
    """PUT /api/workspaces/{name} — update a workspace."""
    from kiro_crew.dashboard.handlers._shared import require_owner_dashboard_request

    # Ahead of the 404: whether a workspace exists is not a non-owner's to learn.
    owner_denied = await require_owner_dashboard_request(request, "workspace.update")
    if owner_denied is not None:
        return owner_denied

    name = request.match_info["name"]
    cfg = KiroCrewConfig.load()
    if name not in cfg.workspaces:
        return web.json_response({"error": f"Workspace '{name}' not found"}, status=404)
    # Default cap: the body is a single directory field.
    body, body_err = await read_bounded_json(request)
    if body_err is not None:
        return body_err
    assert body is not None  # read_bounded_json returns (dict, None) on success
    if "dir" in body:
        new_dir = body["dir"]
        _abs = Path(new_dir).expanduser().is_absolute()
        # Resolved for validation only; is_relative_to + is_sensitive_path guard
        # below reject traversals before the value is stored in config.
        resolved = (  # lgtm[py/path-injection]
            Path(new_dir).expanduser().resolve() if _abs
            else (data_home() / new_dir).resolve()
        )
        if is_sensitive_path(str(resolved)):
            _sel().log_api_access(
                caller=request.get("user", "dashboard"),
                operation="workspace.update",
                outcome="denied",
                source="dashboard",
                resources=name,
            )
            return web.json_response({"error": "Invalid directory path"}, status=400)
        if not _abs and not resolved.is_relative_to(data_home().resolve()):
            _sel().log_api_access(
                caller=request.get("user", "dashboard"),
                operation="workspace.update",
                outcome="denied",
                source="dashboard",
                resources=name,
            )
            return web.json_response({"error": "Invalid directory path"}, status=400)
        if resolved == data_home().resolve():
            _sel().log_api_access(
                caller=request.get("user", "dashboard"),
                operation="workspace.update",
                outcome="denied",
                source="dashboard",
                resources=name,
            )
            return web.json_response(
                {"error": "Cannot use config root as workspace directory"}, status=400
            )
        existing_dirs = {
            (data_home() / ws.dir).resolve()
            if not Path(ws.dir).expanduser().is_absolute()
            else Path(ws.dir).expanduser().resolve()
            for n, ws in cfg.workspaces.items() if n != name
        }
        if resolved in existing_dirs:
            return web.json_response(
                {"error": f"Directory '{new_dir}' is already used by another workspace"},
                status=409,
            )
        cfg.workspaces[name].dir = new_dir
    cfg.save()
    _sel().log_api_access(
        caller=request.get("user", "dashboard"),
        operation="workspace.update",
        outcome="success",
        source="dashboard",
        resources=name,
    )
    return web.json_response({"ok": True, "name": name})


async def api_workspaces_delete(request: web.Request) -> web.Response:
    """DELETE /api/workspaces/{name} — delete a workspace."""
    from kiro_crew.dashboard.handlers._shared import require_owner_dashboard_request

    # Ahead of the 404/409 guards: those are referential, not authorization, and
    # this handler reaches `cfg.save()` with an entry removed.
    owner_denied = await require_owner_dashboard_request(request, "workspace.delete")
    if owner_denied is not None:
        return owner_denied

    name = request.match_info["name"]
    cfg = KiroCrewConfig.load()
    if name not in cfg.workspaces:
        return web.json_response({"error": f"Workspace '{name}' not found"}, status=404)
    if name == cfg.default_workspace:
        return web.json_response(
            {"error": f"Cannot delete default workspace '{name}'. Change default_workspace first."},
            status=409,
        )
    referencing = [a for a, ac in cfg.agents.items() if ac.workspace == name]
    if referencing:
        return web.json_response(
            {"error": f"Workspace '{name}' is referenced by agents: {', '.join(referencing)}"},
            status=409,
        )
    del cfg.workspaces[name]
    cfg.save()
    _sel().log_api_access(
        caller=request.get("user", "dashboard"),
        operation="workspace.delete",
        outcome="success",
        source="dashboard",
        resources=name,
    )
    return web.json_response({"ok": True})


def _validate_dashboard_path(raw: str) -> str | None:
    """Validate a file path through hooks.py enforcement layer."""
    from kiro_crew.hooks import validate_file_path  # noqa: F811

    return validate_file_path(raw)


async def api_file_watch(request: web.Request) -> web.StreamResponse:
    """GET /api/file-watch?path=... — SSE stream of file content changes."""

    raw_path = request.query.get("path", "")
    try:
        validate_tool_args({"path": raw_path}, FILE_READ_SCHEMA)
    except ValidationError:
        _sel().log_tool_invocation(
            session_key="dashboard", tool_name="file_watch", outcome="denied", resources=raw_path
        )
        return web.json_response({"error": "invalid input"}, status=400)

    path = _validate_dashboard_path(raw_path)
    if not path:
        _sel().log_tool_invocation(
            session_key="dashboard", tool_name="file_watch", outcome="denied", resources=raw_path
        )
        return web.json_response({"error": "invalid or forbidden path"}, status=400)

    if not os.path.isfile(path):
        _sel().log_tool_invocation(
            session_key="dashboard", tool_name="file_watch", outcome="not_found", resources=path
        )
        return web.json_response({"error": "not found"}, status=404)

    _sel().log_tool_invocation(
        session_key="dashboard", tool_name="file_watch", outcome="success", resources=path
    )

    resp = web.StreamResponse()
    resp.content_type = "text/event-stream"
    resp.headers["Cache-Control"] = "no-cache"
    resp.headers["X-Accel-Buffering"] = "no"
    await resp.prepare(request)

    poll_interval = 1.0
    read_cap = 512_000
    last_mtime: float = 0.0
    last_content = ""
    resolved_at_start = await asyncio.to_thread(os.path.realpath, path)

    def _read_file(p: str, cap: int) -> str:
        with open(p, "r", encoding="utf-8", errors="replace") as f:
            return f.read(cap)

    try:
        while not (request.transport is None or request.transport.is_closing()):
            try:
                stat = await asyncio.to_thread(os.stat, path)
                mtime = stat.st_mtime
            except FileNotFoundError:
                await asyncio.sleep(poll_interval)
                continue

            if mtime != last_mtime:
                last_mtime = mtime
                current_resolved = await asyncio.to_thread(os.path.realpath, path)
                if current_resolved != resolved_at_start:
                    logger.warning(
                        "file-watch: symlink changed after validation: %s -> %s",
                        resolved_at_start,
                        current_resolved,
                    )
                    _sel().log_tool_invocation(
                        session_key="dashboard",
                        tool_name="file_watch",
                        outcome="denied",
                        resources=path,
                    )
                    break
                try:
                    content = await asyncio.to_thread(_read_file, current_resolved, read_cap)
                    content = redact(content)
                except Exception:
                    logger.warning("file-watch read error for %s", path, exc_info=True)
                    await asyncio.sleep(poll_interval)
                    continue

                if content != last_content:
                    last_content = content
                    # ensure_ascii=False keeps multi-byte content (e.g. CJK)
                    # inspectable as-is in DevTools instead of \uXXXX escapes,
                    # and produces smaller payloads. Body bytes are still
                    # valid UTF-8 because we explicitly .encode() below.
                    payload = json.dumps({"content": content, "mtime": mtime}, ensure_ascii=False)
                    await resp.write(f"data: {payload}\n\n".encode("utf-8"))

            await asyncio.sleep(poll_interval)
    except (ConnectionResetError, asyncio.CancelledError, ClientConnectionResetError):
        pass

    return resp


async def api_file_read(request: web.Request) -> web.Response:
    """GET /api/file-read?path=... — read file content for the markdown panel."""
    from kiro_crew.validation import (  # noqa: F811
        FILE_READ_SCHEMA,
        ValidationError,
        validate_tool_args,
    )

    raw_path = request.query.get("path", "")
    # Resolve relative paths against project dir when resolve=1
    if request.query.get("resolve") == "1":
        raw_path, _resolve_err = _resolve_project_relative(raw_path)
        if _resolve_err == "cannot_resolve":
            return web.json_response(
                {"error": "cannot resolve: no project dir configured"},
                status=400,
            )
        if _resolve_err == "outside_project":
            return web.json_response(
                {"error": "path outside project directory"},
                status=400,
            )

    try:
        validate_tool_args({"path": raw_path}, FILE_READ_SCHEMA)
    except ValidationError:
        _sel().log_tool_invocation(
            session_key="dashboard",
            tool_name="file_read",
            outcome="denied",
            resources=raw_path,
        )
        return web.json_response({"error": "invalid input"}, status=400)

    path = _validate_dashboard_path(raw_path)
    if not path:
        _sel().log_tool_invocation(
            session_key="dashboard",
            tool_name="file_read",
            outcome="denied",
            resources=raw_path,
        )
        return web.json_response({"error": "invalid or forbidden path"}, status=400)
    if not os.path.isfile(path):
        # Both a directory and a missing path are 404 for a READ — there is no
        # file content to return either way — but the caller needs to tell them
        # apart. The dashboard renders a markdown path chip as a folder
        # affordance when the path is a directory and suppresses the chip
        # entirely when the path is not on disk; without this header both look
        # like "file not found", which is actively wrong for a directory.
        #
        # Sitting ahead of the HEAD branch below, one probe covers GET and HEAD.
        # `path` is already realpath-canonical and denylist-checked here, so
        # isdir() discloses nothing that the status code did not already.
        is_dir = os.path.isdir(path)
        _sel().log_tool_invocation(
            session_key="dashboard", tool_name="file_read", outcome="not_found", resources=path
        )
        return web.json_response(
            {"error": "is a directory" if is_dir else "not found"},
            status=404,
            headers={"X-Path-Kind": "dir" if is_dir else "missing"},
        )
    if request.method == "HEAD":
        _sel().log_tool_invocation(
            session_key="dashboard", tool_name="file_read", outcome="success", resources=path
        )
        return web.Response(status=200, headers={"X-Path-Kind": "file"})
    try:
        read_cap = 512_000
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            content = f.read(read_cap + 1)
        truncated = len(content) > read_cap
        content = content[:read_cap]
        content = redact(content)
        _sel().log_tool_invocation(
            session_key="dashboard", tool_name="file_read", outcome="success", resources=path
        )
        headers = {"X-Truncated": "true"} if truncated else {}
        # Pick a sensible content_type per file extension so browsers and
        # debuggers (DevTools "Response" preview, curl) interpret the body
        # correctly. JSON files in particular benefit from application/json
        # so DevTools renders the body as a tree instead of raw text.
        # aiohttp appends "; charset=utf-8" automatically when text= is set.
        #
        # Security: HTML files are deliberately served as text/plain to
        # prevent stored-XSS via <script> tags or on* attribute handlers in
        # user/LLM-generated content. The dashboard's HtmlViewer renders
        # HTML files via a sandboxed srcDoc iframe, so the file-read
        # endpoint never needs to deliver executable HTML.
        ext = os.path.splitext(path)[1].lower()
        if ext == ".json":
            ct = "application/json"
        elif ext == ".jsonl":
            # JSONL (newline-delimited JSON) is NOT a valid JSON document —
            # the registered MIME type is application/x-ndjson. Serving it
            # as application/json would make DevTools / JsonViewer try to
            # parse the whole body as one JSON value and fail.
            ct = "application/x-ndjson"
        elif ext == ".csv":
            ct = "text/csv"
        elif ext in (".md", ".markdown"):
            ct = "text/markdown"
        else:
            ct = "text/plain"
        return web.Response(text=content, content_type=ct, headers=headers)
    except Exception:
        logging.getLogger(__name__).exception("file_read failed for %s", path)
        _sel().log_tool_invocation(
            session_key="dashboard", tool_name="file_read", outcome="failure", resources=path
        )
        return web.json_response({"error": "failed to read file"}, status=500)


class _OpenDenied(NamedTuple):
    """A refusal from :func:`_open_checked_file`: why, and the path to log.

    ``code`` uses the machine vocabulary the streaming endpoint already
    exposes (``invalid_path`` / ``sensitive_path`` / ``not_found`` /
    ``symlink_refused`` / ``file_too_large`` / ``read_failed``); each adopter
    maps it onto its own SEL outcome and response body, which is where the
    endpoints legitimately differ. ``path`` is the raw input for
    ``invalid_path`` (validation produced nothing) and the validated path
    otherwise -- exactly what each adopter logs today.
    """

    code: str
    path: str


class _CheckedFile(NamedTuple):
    """A successful :func:`_open_checked_file`: the checked open file.

    ``size`` is the fstat size of THIS fd -- authoritative for a streaming
    caller that must announce a length, advisory for whole-read callers
    whose bounded read is their own size guard.
    """

    path: str
    file: BinaryIO
    size: int


def _open_checked_file(
    raw_path: str,
    *,
    tool_name: str,
    fstat_cap: int | None = None,
    log_open_failure: bool = True,
) -> _CheckedFile | _OpenDenied:
    """The open-and-check half of the file-serving security prefix.

    validate -> sensitive-path check -> is-file -> ``_open_rb_nofollow`` ->
    fstat (cap enforced only when *fstat_cap* is passed), then RETURNS the
    checked open file object. What happens to the bytes afterwards is
    per-endpoint POLICY and stays with the caller: the whole-read envelope
    (:func:`_open_checked`) reads and closes it, the streaming endpoint
    sniffs and serves ranges from it, the sheet endpoint hands it to the
    workbook parser. The split exists because the streaming and sheet
    endpoints must keep the open file object, so they cannot use the
    whole-read envelope -- sharing the prefix keeps ONE copy of this
    boundary for every endpoint.

    The sensitive-path gate runs through this module's import-time
    ``is_sensitive_path`` alias -- ONE binding for one guard, so a test
    override (or a future hardening change) applied to
    ``files.is_sensitive_path`` is observed by every adopter instead of
    landing on whichever binding an endpoint happened to import.

    *fstat_cap* is the streaming endpoint's size policy: its fd stays open
    for range reads, so the announced size must be authoritative up front.
    Whole-read adopters pass no cap here -- an fstat pre-check races a
    concurrent writer (the file can grow between the stat and the read),
    while their bounded read caps memory unconditionally.

    *log_open_failure* keeps log volume a per-endpoint decision: a
    caller-reachable open failure (mode-000 file, EACCES on a parent) writes
    a full traceback per request when True. The whole-read envelope wants
    that traceback (its 500 is the only signal); the streaming and sheet
    endpoints answer a coded refusal instead and pass False, so a request
    loop against a known-unreadable path cannot amplify into the log.

    Synchronous by design -- callers run it on a worker thread. Refusals are
    returned as typed codes, not responses: SEL logging and the HTTP body
    vocabulary belong to each endpoint.
    """
    import kiro_crew.dashboard.handlers as _h  # noqa: F811  # circular import

    try:
        path = _h._validate_dashboard_path(raw_path)
    except ValueError:
        # A malformed path (an embedded NUL makes realpath raise) is an
        # invalid path, not a crash.
        path = None
    if not path:
        return _OpenDenied("invalid_path", raw_path)
    if is_sensitive_path(path):
        return _OpenDenied("sensitive_path", path)
    if not os.path.isfile(path):
        return _OpenDenied("not_found", path)
    # Symlinks rejected atomically (O_NOFOLLOW on POSIX; lstat guard +
    # O_BINARY on Windows -- see _open_rb_nofollow).
    try:
        fd = _open_rb_nofollow(path)
    except OSError as exc:
        if exc.errno == errno.ELOOP:  # symlink with O_NOFOLLOW
            return _OpenDenied("symlink_refused", path)
        if log_open_failure:
            # The only traceback for a failed open: adopters map the code
            # onto an outcome, and SEL records outcome, not cause.
            logger.exception("%s open failed for %s", tool_name, path)
        return _OpenDenied("read_failed", path)
    fobj = os.fdopen(fd, "rb")
    try:
        # fstat is authoritative for THIS fd; a file that grows afterwards
        # only extends past the size announced here, never past the cap.
        size = os.fstat(fobj.fileno()).st_size
    except OSError:
        with contextlib.suppress(Exception):
            fobj.close()
        if log_open_failure:
            logger.exception("%s fstat failed for %s", tool_name, path)
        return _OpenDenied("read_failed", path)
    if fstat_cap is not None and size > fstat_cap:
        fobj.close()
        return _OpenDenied("file_too_large", path)
    return _CheckedFile(path=path, file=fobj, size=size)


class _OpenRefusal(NamedTuple):
    """A refusal from :func:`_open_checked`: the response to return, already audited."""

    response: web.Response


class _OpenedFile(NamedTuple):
    """A successful :func:`_open_checked`: the validated path and full bytes."""

    path: str
    data: bytes


def _open_checked(
    raw_path: str,
    *,
    tool_name: str,
    max_bytes: int,
) -> _OpenedFile | _OpenRefusal:
    """The dashboard file endpoints' shared WHOLE-READ envelope.

    The open-and-check half lives in :func:`_open_checked_file` (the prefix
    shared with the streaming and sheet endpoints); this layer is the
    whole-read policy on top: bounded read (cap enforced on the bytes
    actually read, so a concurrent writer cannot outgrow it), close, and the
    mapping of every refusal onto this envelope's SEL vocabulary and
    response bodies.

    This is a SECURITY boundary: hand-rolled copies of one mean a future
    hardening fix — a new TOCTOU guard, a tightened sniff, a cap change —
    lands in some and silently leaves the others on the old posture (the
    same shape the zip-vetting surfaces guard against).

    Per-endpoint POLICY stays with the endpoint and is passed in rather than
    copied: which cap applies, and what the endpoint does with the bytes
    afterwards (``data`` is the full file — a caller sniffing magic slices
    its own header). Only the envelope is shared.

    Returns the opened result, or a refusal carrying the response to return —
    a typed either, so a caller cannot accidentally use the data on a refusal
    path the way an ``(data, error)`` tuple invites.

    Synchronous by design — callers offload it via ``asyncio.to_thread``:
    everything here is blocking file I/O and must not run on the event loop.
    SEL audit writes are thread-safe (locked appends).
    """

    def _log(outcome: str, res: str, error: str = "") -> None:
        kw = {"error": error} if error else {}
        _sel().log_tool_invocation(
            session_key="dashboard", tool_name=tool_name,
            outcome=outcome, resources=res, **kw,
        )

    checked = _open_checked_file(raw_path, tool_name=tool_name)
    if isinstance(checked, _OpenDenied):
        code, res = checked.code, checked.path
        if code == "invalid_path":
            _log("denied", res)
            return _OpenRefusal(
                web.json_response({"error": "invalid or forbidden path"}, status=400)
            )
        if code == "sensitive_path":
            _log("denied", res, "sensitive_path")
            return _OpenRefusal(
                web.json_response({"error": "sensitive path blocked"}, status=403)
            )
        if code == "not_found":
            _log("not_found", res)
            return _OpenRefusal(web.json_response({"error": "not found"}, status=404))
        if code == "symlink_refused":
            _log("denied", res, "symlink_rejected")
            return _OpenRefusal(
                web.json_response({"error": "symlinks not allowed"}, status=403)
            )
        if code == "file_too_large":
            # Reachable only through a caller that passes fstat_cap; mapped so
            # a policy refusal can never masquerade as the 500 below.
            _log("denied", res, "file_too_large")
            return _OpenRefusal(
                web.json_response(
                    {"error": "file too large", "code": "file_too_large"}, status=413
                )
            )
        # read_failed: the residual code. (This envelope's own size guard is
        # the bounded read below, because an fstat pre-check races a
        # concurrent writer while reading at most cap+1 bytes bounds memory
        # unconditionally -- the same shape as _load_sheet_payload's guard.)
        _log("failure", res)
        return _OpenRefusal(
            web.json_response({"error": "cannot read file", "code": "read_failed"}, status=500)
        )

    path = checked.path
    try:
        with checked.file as f:
            data = f.read(max_bytes + 1)
    except OSError:
        # Keep the traceback for a failed read: this 500 is the only signal,
        # and SEL records outcome, not cause.
        logger.exception("%s read failed for %s", tool_name, path)
        _log("failure", path)
        return _OpenRefusal(web.json_response({"error": "cannot read file"}, status=500))
    if len(data) > max_bytes:
        _log("denied", path, "file_too_large")
        return _OpenRefusal(
            web.json_response({"error": "file too large"}, status=413)
        )

    return _OpenedFile(path=path, data=data)


async def api_file_download(request: web.Request) -> web.Response:
    """GET /api/file-download?path=... — download a file as raw bytes.

    Sibling of /api/file-read. file-read decodes content as UTF-8 with
    errors='replace' to render text in the markdown panel; that mode
    corrupts binary files (.docx, .pdf, images) by replacing non-text
    bytes with U+FFFD. This endpoint streams the original bytes, sets
    Content-Disposition: attachment, and applies X-Content-Type-Options:
    nosniff to keep the browser from rendering the response inline.

    Security: same path-validation as file-read (validate_tool_args,
    _validate_dashboard_path, sensitive-path filter). Symlinks rejected
    via O_NOFOLLOW. Files larger than _MAX_UPLOAD_BYTES are rejected.
    Text files are still scanned for sensitive content (credentials and
    exfiltration URLs); a positive hit aborts the download. Binary
    files are served as-is without a MIME allowlist, since attachment
    disposition + nosniff prevents inline rendering on the dashboard
    origin.
    """
    # Path validation now happens inside ``_open_checked``, which keeps the
    # late-binding ``handlers`` alias so tests can still monkey-patch
    # ``_validate_dashboard_path`` (legitimate circular-import workaround,
    # listed as an exception in the top-level-imports rule).
    raw_path = request.query.get("path", "")
    # Resolve relative paths against project dir when resolve=1 (mirrors api_file_read)
    if request.query.get("resolve") == "1":
        raw_path, _resolve_err = _resolve_project_relative(raw_path)
        if _resolve_err == "cannot_resolve":
            return web.json_response(
                {"error": "cannot resolve: no project dir configured"}, status=400,
            )
        if _resolve_err == "outside_project":
            return web.json_response(
                {"error": "path outside project directory"}, status=400,
            )

    try:
        validate_tool_args({"path": raw_path}, FILE_READ_SCHEMA)
    except ValidationError:
        _sel().log_tool_invocation(
            session_key="dashboard", tool_name="file_download",
            outcome="denied", resources=raw_path,
        )
        return web.json_response({"error": "invalid input"}, status=400)

    # Envelope shared with api_file_raw (#4031). No header sniff: this endpoint
    # serves attachment + nosniff rather than choosing a content type. Offloaded
    # to a worker thread: the envelope is synchronous file I/O (realpath, open,
    # fstat, full read up to the cap) and must not block the event loop.
    opened = await asyncio.to_thread(
        _open_checked, raw_path, tool_name="file_download", max_bytes=_MAX_UPLOAD_BYTES,
    )
    if isinstance(opened, _OpenRefusal):
        return opened.response
    path, data = opened.path, opened.data

    # Defense in depth: scan content for credentials / exfil URLs via the
    # context-aware redact() shim, which runs BOTH the exfil-URL and credential
    # passes (exfil URLs first so embedded credentials in URL fragments are
    # caught) and additionally applies a loaded companion's extra regexes before
    # content reaches an external surface.
    #
    # Mostly-binary files can still hide credential patterns in their
    # decodable runs (e.g. an ASCII-art `AKIA...` with one stray non-UTF-8
    # byte). Decoding with errors='replace' for the *scan only* (the served
    # bytes are still raw) ensures the credential pass cannot be bypassed
    # by sprinkling a single non-UTF-8 byte into the file.
    try:
        text = data.decode("utf-8")
    except UnicodeDecodeError:
        text = data.decode("utf-8", errors="replace")
    # Route through the context-aware redact() so a loaded companion's extra
    # credential regexes also abort the download; the scrubbed != text diff is
    # the gate (no count needed).
    scrubbed = redact(text)
    if scrubbed != text:
        _sel().log_tool_invocation(
            session_key="dashboard", tool_name="file_download",
            outcome="denied", resources=path, error="content_redacted",
        )
        return web.json_response(
            {"error": "file content was redacted; download aborted"}, status=400,
        )

    safe_name = urllib.parse.quote(os.path.basename(path), safe="")
    content_type, _ = mimetypes.guess_type(path)
    if not content_type:
        content_type = "application/octet-stream"

    _sel().log_tool_invocation(
        session_key="dashboard", tool_name="file_download",
        outcome="success", resources=path,
    )
    return web.Response(
        body=data,
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{safe_name}",
            "Content-Type": content_type,
            "X-Content-Type-Options": "nosniff",
        },
    )


# Extensions previewable via kiro_crew.doc_parser (OOXML docx/pptx). Legacy
# binary formats (.doc, .ppt), the OpenDocument family (.odt/.ods/.odp), and
# spreadsheet formats (.xls/.xlsx) fall through to the download card because
# doc_parser only understands ZIP+XML OOXML, and adding openpyxl or a legacy
# OLE reader would grow the dependency tree noticeably for a preview feature.
_OFFICE_PREVIEWABLE_EXT = {".docx", ".pptx"}
# Cap the returned text so a huge .docx doesn't blow the JSON payload / DOM.
# Mirrors api_file_read's 512 KB read cap. Anything larger is truncated and
# the frontend shows a "Download for full contents" affordance.
_OFFICE_PREVIEW_CAP = 512_000


class _PreviewUnsupported(Exception):
    """The validated path's extension is outside :data:`_OFFICE_PREVIEWABLE_EXT`.

    Endpoint-local, mirroring :class:`_SheetRefusal`: ``_OpenDenied``'s codes
    are the SHARED file-serving boundary's vocabulary, and this is this
    endpoint's own FORMAT policy rather than a security refusal, so it does
    not belong in that enum. Raised from inside the worker callback so the
    checked file object is closed by its ``with`` block on the same thread.
    """


async def api_file_office_preview(request: web.Request) -> web.Response:
    """GET /api/file-office-preview?path=... — extract inline text preview from a .docx/.pptx.

    Sibling of /api/file-download. file-download streams original bytes for
    saving to disk; this endpoint returns plaintext extracted from the
    OOXML XML inside so the dashboard can render a scrollable preview of
    the document contents in place of the "can't view a binary" download
    card — a common ask for anyone browsing shared reports in the file
    tree without wanting to save each one.

    Uses ``kiro_crew.doc_parser.extract_text`` which parses the .docx /
    .pptx ZIP+XML with hardened defusedxml (XXE-safe) and returns "" on
    any failure. python-docx / python-pptx are not required.

    Not supported (fall through to download): .doc, .ppt, .xls, .xlsx,
    .odt, .ods, .odp. The frontend keeps the download card for these.

    Security: the open-and-check prefix is the SHARED
    :func:`_open_checked_file` (dashboard path validation, sensitive-path
    block, is-file, symlink-refusing ``_open_rb_nofollow`` — atomic
    O_NOFOLLOW on POSIX, lstat guard on Windows — then fstat), never a
    hand-rolled second spelling of it, so a future hardening change to that
    boundary lands here too. This endpoint's own POLICY on top is the 50 MB
    ``fstat_cap``, the ``.docx``/``.pptx`` format gate, the aggregate
    extraction budget, and credential redaction before the preview cap is
    applied. All of it — validation, open, fstat, ZIP+XML parsing,
    redaction — runs in ONE worker-thread hop, like ``api_file_sheet``.
    """
    raw_path = request.query.get("path", "")

    def _log(outcome: str, res: str, error: str = "") -> None:
        kw = {"error": error} if error else {}
        _sel().log_tool_invocation(
            session_key="dashboard", tool_name="file_office_preview",
            outcome=outcome, resources=res, **kw,
        )

    # Resolve relative paths against project dir when resolve=1. Uses the
    # shared helper (same as api_file_read / api_file_download / file-raw):
    # it passes Windows-absolute/UNC shapes through to the validator, whose
    # network-path gate runs BEFORE realpath — never re-implement this inline.
    if request.query.get("resolve") == "1":
        raw_path, _resolve_err = _resolve_project_relative(raw_path)
        if _resolve_err == "cannot_resolve":
            _log("denied", request.query.get("path", ""), "cannot_resolve")
            return web.json_response(
                {"error": "cannot resolve: no project dir configured", "code": "no_project_dir"},
                status=400,
            )
        if _resolve_err == "outside_project":
            _log("denied", request.query.get("path", ""), "outside_project")
            return web.json_response(
                {"error": "path outside project directory", "code": "path_outside_project"},
                status=400,
            )

    try:
        validate_tool_args({"path": raw_path}, FILE_READ_SCHEMA)
    except ValidationError:
        _log("denied", raw_path)
        return web.json_response({"error": "invalid input", "code": "invalid_input"}, status=400)

    # The validated path once the shared prefix produces one -- exported by
    # the worker callback so the exception handlers log the same SEL resource
    # the success path does.
    res_path = raw_path

    def _open_and_extract() -> dict[str, object] | _OpenDenied:
        """Open-and-check plus extract, in ONE worker-thread hop.

        Everything here is blocking I/O or CPU-bound — realpath validation,
        the sensitive-path screen, the open, the fstat, ZIP decompression,
        XML parsing, redaction — so none of it may run on the event loop: an
        NFS/FUSE-backed document makes even the validate/open envelope block
        for seconds, stalling every session's streaming and the liveness
        heartbeat.

        The checked open file object never crosses back to the event loop:
        every path that opens it also closes it on THIS thread (refusals
        close inside the prefix; the ``with`` block below covers the rest,
        the format refusal included). A cancellation of the awaiting task
        therefore cannot strand an open file in a discarded future or
        finalize one on the loop — the future's result is only ever a
        payload dict or a typed refusal.
        """
        nonlocal res_path
        # fstat_cap is this endpoint's size gate, enforced on the fd BEFORE
        # any ZIP parsing: zipfile.ZipFile materializes the archive's central
        # directory in memory, bounded only by the file itself, so a crafted
        # archive could otherwise exhaust memory before doc_parser's
        # per-entry and aggregate budgets ever apply. Same 50 MB ceiling as
        # file uploads. log_open_failure=False: this endpoint answers a coded
        # refusal, so a request loop against a known-unreadable path cannot
        # amplify into the log.
        checked = _open_checked_file(
            raw_path,
            tool_name="file_office_preview",
            fstat_cap=_MAX_UPLOAD_BYTES,
            log_open_failure=False,
        )
        if isinstance(checked, _OpenDenied):
            return checked
        res_path = checked.path
        with checked.file as fobj:
            if os.path.splitext(checked.path)[1].lower() not in _OFFICE_PREVIEWABLE_EXT:
                raise _PreviewUnsupported(checked.path)
            # extract_text parses through the SAME handle the prefix opened
            # and fstat-ed (its opt-in fileobj parameter), so the bytes
            # parsed are exactly the bytes measured — no stat→open TOCTOU
            # window. max_chars bounds AGGREGATE extraction (cap + 1 keeps
            # the truncation flag detectable): a deck with thousands of
            # slides stops parsing at the budget instead of accumulating
            # unbounded text. It never raises — returns "" on any failure.
            text = extract_text(
                checked.path,
                filename=os.path.basename(checked.path),
                max_chars=_OFFICE_PREVIEW_CAP + 1,
                fileobj=fobj,
            )
        truncated = len(text) > _OFFICE_PREVIEW_CAP
        # Redact BEFORE truncating: slicing first could cut a credential
        # across the cap boundary, leaving an unmatched prefix the redactor
        # no longer recognizes. Redaction may change the length, so the
        # truncation flag is computed from the raw extraction above.
        text = redact(text)
        if truncated:
            text = text[:_OFFICE_PREVIEW_CAP]
        return {
            "text": text,
            "truncated": truncated,
            # No `empty` field: doc_parser returns "" for both a genuinely
            # blank document and a parse failure, so the two are
            # indistinguishable here. The frontend treats empty `text` as
            # "no preview available" and falls back to the download card.
        }

    try:
        result = await asyncio.to_thread(_open_and_extract)
    except asyncio.CancelledError:
        # Gateway shutdown / client disconnect while the worker thread is
        # parsing: the access attempt already happened, so record it before
        # propagating — CancelledError is a BaseException and would bypass
        # the Exception handler below, leaving the access unaudited. No
        # resource handling here: the worker callback owns the file's whole
        # lifetime.
        _log("cancelled", res_path)
        raise
    except _PreviewUnsupported:
        # 415 (not 400) so the frontend can distinguish "unsupported format,
        # keep showing the download card" from "invalid input, something's
        # actually wrong". The frontend short-circuits known-unsupported
        # extensions client-side, so this branch is the safety net (direct
        # API calls, frontend/backend list drift).
        _log("denied", res_path, "unsupported_preview_format")
        return web.json_response(
            {
                "error": "unsupported format for inline preview",
                "code": "unsupported_preview_format",
            },
            status=415,
        )
    except Exception:  # noqa: BLE001  # last-resort guard; doc_parser already logs
        logger.exception("file_office_preview extract_text failed for %s", res_path)
        _log("failure", res_path)
        return web.json_response(
            {"error": "failed to extract preview", "code": "preview_extraction_failed"},
            status=500,
        )
    if isinstance(result, _OpenDenied):
        # The shared prefix's typed refusals, mapped onto this endpoint's SEL
        # outcomes and response vocabulary — the part that legitimately
        # differs per endpoint.
        code, res = result.code, result.path
        if code == "invalid_path":
            _log("denied", res)
            return web.json_response(
                {"error": "invalid or forbidden path", "code": "forbidden_path"}, status=400,
            )
        if code == "sensitive_path":
            _log("denied", res, "sensitive_path")
            return web.json_response(
                {"error": "sensitive path blocked", "code": "sensitive_path"}, status=403,
            )
        if code == "not_found":
            _log("not_found", res)
            return web.json_response({"error": "not found", "code": "not_found"}, status=404)
        if code == "symlink_refused":
            _log("denied", res, "symlink_rejected")
            return web.json_response(
                {"error": "symlinks not allowed", "code": "symlink_rejected"}, status=403,
            )
        if code == "file_too_large":
            _log("denied", res, "file_too_large")
            return web.json_response(
                {
                    "error": (
                        "file too large for preview "
                        f"(max {_MAX_UPLOAD_BYTES // 1024 // 1024}MB)"
                    ),
                    "code": "file_too_large",
                },
                status=413,
            )
        # read_failed: the residual code.
        _log("failure", res)
        return web.json_response(
            {"error": "cannot read file", "code": "file_read_failed"}, status=500,
        )
    _log("success", res_path)
    return web.json_response(result)


async def api_file_raw(request: web.Request) -> web.Response:
    """GET /api/file-raw?path=... — serve a file with its native content type (images, etc.)."""
    # Envelope (validate -> sensitive -> nofollow-open -> bounded read) is
    # shared with api_file_download so a hardening change lands on both (#4031).
    # Offloaded to a worker thread: the envelope is synchronous file I/O and
    # must not block the event loop (same shape as api_file_stream's _open_media).
    opened = await asyncio.to_thread(
        _open_checked,
        request.query.get("path", ""),
        tool_name="file_raw",
        max_bytes=_MAX_UPLOAD_BYTES,
    )
    if isinstance(opened, _OpenRefusal):
        return opened.response
    path, data = opened.path, opened.data
    # SNIFF_BYTES: the shared raster sniffer's documented minimum, and enough
    # for every magic matched below (WebP's form tag ends at byte 12).
    header = data[:SNIFF_BYTES]

    def _log(outcome: str, res: str) -> None:
        _sel().log_tool_invocation(
            session_key="dashboard", tool_name="file_raw", outcome=outcome, resources=res,
        )

    # Raster types are detected by the shared sniffer
    # (kiro_crew.messaging.raster), which requires the full PNG signature and
    # WebP's form tag at offset 8 — so a RIFF/WAVE audio file is not served as
    # an image. TIFF and ICO keep local rows (_READ_PATH_EXTRA_MAGIC).
    content_type = sniff_raster_mime(header)
    if content_type is None:
        for magic, mime in _READ_PATH_EXTRA_MAGIC:
            if header.startswith(magic):
                content_type = mime
                break
    # SVG: XML-based, no magic bytes
    if not content_type:
        stripped = data.lstrip(b"\xef\xbb\xbf").lstrip()
        if stripped.startswith(b"<svg") or (
            stripped.startswith(b"<?xml") and b"<svg" in data[:4096]
        ):
            content_type = "image/svg+xml"
    # PDF: %PDF magic bytes
    if not content_type:
        if header.startswith(b"%PDF"):
            content_type = "application/pdf"
    if not content_type:
        _log("denied", path)
        return web.json_response({"error": "file content is not a recognized format"}, status=403)
    _log("success", path)
    headers = {"Content-Type": content_type, "X-Content-Type-Options": "nosniff"}
    if content_type == "image/svg+xml":
        headers["Content-Security-Policy"] = "script-src 'none'; style-src 'unsafe-inline'"
    return web.Response(body=data, headers=headers)


# ── /api/file-stream: Range-capable audio/video serving ─────────────────────
# The media cap is deliberately larger than _MAX_UPLOAD_BYTES: screen
# recordings routinely exceed 50 MB, and unlike file-raw this endpoint never
# materializes the file in memory -- Range streaming reads bounded chunks, so
# the cap only bounds what one URL can address, not per-request memory.
_STREAM_MAX_BYTES = 2 * 1024 * 1024 * 1024
_STREAM_CHUNK_BYTES = 256 * 1024
# Text-exfiltration probe window. Real media is binary within the first
# bytes; content that decodes as UTF-8 text this deep is a text file wearing
# a media magic, which the redaction scan below must see.
_STREAM_TEXT_PROBE_BYTES = 64 * 1024


def _resolve_project_relative(raw: str) -> tuple[str, str | None]:
    """Resolve a relative path against KIROCREW_PROJECT_DIR (resolve=1).

    Returns (path, None) on success -- absolute and ~-paths pass through
    unchanged -- or ("", error_code) with "cannot_resolve" (no project dir
    configured) or "outside_project" (the joined path escapes the project
    directory after realpath).
    """
    if not raw or raw.startswith(("/", "~")):
        return raw, None
    # Windows-absolute shapes (UNC \\server\share, drive C:\...) are not
    # project-relative: pass them to the validator unchanged. Joining them
    # would let os.path.realpath contact the named host (SMB round-trip)
    # before any validation runs; the validator's own network-path gate
    # sits BEFORE its realpath, so it is the safe place for these.
    if raw.startswith("\\") or ntpath.splitdrive(raw)[0]:
        return raw, None
    proj = os.environ.get("KIROCREW_PROJECT_DIR", "")
    if not proj:
        return "", "cannot_resolve"
    candidate = os.path.realpath(os.path.join(proj, raw))
    resolved_proj = os.path.realpath(proj)
    if not (candidate == resolved_proj or candidate.startswith(resolved_proj + os.sep)):
        return "", "outside_project"
    return candidate, None


# Container signature -> Content-Type. Sniffed from the file's first bytes so
# the endpoint serves media by CONTENT, not by extension claim (CWE-434 shape,
# same posture as file-raw's image allowlist). Entries are (offset, magic,
# mime). MP4-family uses the ftyp box at offset 4 (bytes 0-3 are the box
# size); WebM and Matroska share the EBML magic and both play in <video>.
_MEDIA_MAGIC: tuple[tuple[int, bytes, str], ...] = (
    (4, b"ftyp", "video/mp4"),          # mp4 / m4v / m4a / mov (BMFF family)
    (0, b"\x1a\x45\xdf\xa3", "video/webm"),  # webm / mkv (EBML)
    (0, b"OggS", "audio/ogg"),          # ogg audio or video; <audio>/<video> both accept
    (0, b"fLaC", "audio/flac"),
    (0, b"ID3", "audio/mpeg"),          # mp3 with ID3v2 tag
    (0, b"\xff\xfb", "audio/mpeg"),     # bare mp3 frame sync (MPEG1 layer3)
    (0, b"\xff\xf3", "audio/mpeg"),
    (0, b"\xff\xf2", "audio/mpeg"),
)


def _sniff_media_type(header: bytes) -> str | None:
    """Return the media Content-Type for ``header`` bytes, or None."""
    for offset, magic, mime in _MEDIA_MAGIC:
        if header[offset:offset + len(magic)] == magic:
            return mime
    # WAV: RIFF....WAVE compound signature (offset 8 discriminates from WebP)
    if header[:4] == b"RIFF" and header[8:12] == b"WAVE":
        return "audio/wav"
    return None


def _parse_range_header(value: str, size: int) -> tuple[int, int] | None:
    """Parse a single-range ``bytes=`` header against ``size``.

    Returns (start, end) inclusive, or None for an unsatisfiable or
    malformed header. Multi-range requests are treated as malformed --
    <audio>/<video> elements only ever issue single ranges, and multipart
    responses would complicate the reader for no consumer.
    """
    if not value.startswith("bytes="):
        return None
    spec = value[len("bytes="):]
    if "," in spec or "-" not in spec:
        return None
    start_s, _, end_s = spec.partition("-")
    try:
        if start_s == "":
            # suffix form: last N bytes
            suffix = int(end_s)
            if suffix <= 0:
                return None
            start = max(0, size - suffix)
            end = size - 1
        else:
            start = int(start_s)
            end = int(end_s) if end_s else size - 1
    except ValueError:
        return None
    if start < 0 or start >= size or end < start:
        return None
    return start, min(end, size - 1)


async def api_file_stream(request: web.Request) -> web.StreamResponse:
    """GET /api/file-stream?path=... -- serve audio/video with Range support.

    Powers inline <video>/<audio> playback in the file viewer. file-raw is
    unsuitable for media: it whole-reads the file into memory, rejects
    anything over the upload cap, and ignores Range headers -- and seeking in
    a media element requires 206 Partial Content. This endpoint follows the
    same security pattern (dashboard path validation, sensitive-path block,
    symlink-refusing open, content sniffing before serving) but streams
    bounded chunks off the event loop, so memory stays constant regardless
    of file size. All reads go through the SAME fd the header was sniffed
    from, so the served bytes cannot be swapped after the check.

    Accepted gap (documented, not a defect): the redaction probe covers the
    first 64 KiB. Complete coverage is unreachable for a Range endpoint --
    the client controls byte offsets, so any pattern scan can be split
    across range boundaries -- and the sibling binary-serving endpoint
    (file-raw) performs no content scan at all. The probe exists to catch
    the honest-mistake shape: a text file wearing a forged media magic.
    """

    def _log(outcome: str, res: str) -> None:
        _sel().log_tool_invocation(
            session_key="dashboard", tool_name="file_stream", outcome=outcome, resources=res,
        )

    raw_path = request.query.get("path", "")
    resolve_requested = request.query.get("resolve") == "1"

    def _open_media(raw: str) -> tuple:
        """Validate, open, and sniff the media file. Runs on a worker thread.

        The open-and-check prefix is the shared :func:`_open_checked_file`
        (validate -> sensitive-path -> nofollow open -> fstat cap), with this
        endpoint's stream cap passed in as policy; what stays here is the
        endpoint's own policy: relative-path resolution against the project
        dir (the resolve=1 contract shared with file-read/file-download), the
        media sniff, and the text probe. Returns either
        ("ok", file_object, size, content_type, path) or a refusal tuple
        ("refused", code, path_for_log).
        """
        if resolve_requested:
            try:
                raw, resolve_err = _resolve_project_relative(raw)
            except ValueError:
                # A malformed path (embedded NUL) is an invalid path, not a
                # crash -- same verdict the shared prefix gives one.
                return ("refused", "invalid_path", raw)
            if resolve_err:
                return ("refused", resolve_err, raw)
        checked = _open_checked_file(
            raw, tool_name="file_stream", fstat_cap=_STREAM_MAX_BYTES,
            log_open_failure=False,
        )
        if isinstance(checked, _OpenDenied):
            return ("refused", checked.code, checked.path)
        fobj, size, validated = checked.file, checked.size, checked.path
        try:
            header = fobj.read(16)
            content_type = _sniff_media_type(header)
            if not content_type:
                fobj.close()
                return ("refused", "not_media", validated)
            # Sibling-control parity: file-download refuses text content that
            # redact() flags. Media magics can be weak (the bare mp3 frame
            # sync is two bytes), so a credential-bearing TEXT file with a
            # forged prefix must not stream out here. Decode with
            # errors="replace" -- exactly as the download scan does -- so an
            # invalid byte (including the forged magic itself) cannot skip
            # the credential pass; replacement chars break no real credential
            # pattern, and genuine binary media decodes to replacement-dense
            # junk that redact() leaves unchanged. The scan is bounded to the
            # probe window; the full-file scan remains the download path's.
            probe = header + fobj.read(_STREAM_TEXT_PROBE_BYTES - len(header))
            probe_text = probe.decode("utf-8", errors="replace")
            if redact(probe_text) != probe_text:
                fobj.close()
                return ("refused", "content_redacted", validated)
            fobj.seek(0)
        except Exception:
            with contextlib.suppress(Exception):
                fobj.close()
            return ("refused", "read_failed", validated)
        return ("ok", fobj, size, content_type, validated)

    result = await asyncio.to_thread(_open_media, raw_path)
    if result[0] == "refused":
        _, code, res = result
        if code == "not_found":
            outcome = "not_found"
        elif code == "read_failed":
            outcome = "failure"
        else:
            outcome = "denied"
        _log(outcome, res)
        # One literal response per refusal class: the error-response contract
        # requires the {"error", "code"} body and the status to be statically
        # checkable at each call site.
        if code == "invalid_path":
            return web.json_response(
                {"error": "invalid or forbidden path", "code": "invalid_path"}, status=400
            )
        if code == "cannot_resolve":
            return web.json_response(
                {"error": "cannot resolve: no project dir configured", "code": "cannot_resolve"},
                status=400,
            )
        if code == "outside_project":
            return web.json_response(
                {"error": "path outside project directory", "code": "outside_project"},
                status=400,
            )
        if code == "sensitive_path":
            return web.json_response(
                {"error": "sensitive path blocked", "code": "sensitive_path"}, status=403
            )
        if code == "not_found":
            return web.json_response({"error": "not found", "code": "not_found"}, status=404)
        if code == "symlink_refused":
            return web.json_response(
                {"error": "symlinks not allowed", "code": "symlink_refused"}, status=403
            )
        if code == "file_too_large":
            return web.json_response(
                {"error": "file too large", "code": "file_too_large"}, status=413
            )
        if code == "not_media":
            return web.json_response(
                {"error": "file content is not a supported media format", "code": "not_media"},
                status=415,
            )
        if code == "content_redacted":
            return web.json_response(
                {"error": "file content was redacted; stream aborted",
                 "code": "content_redacted"},
                status=400,
            )
        return web.json_response(
            {"error": "cannot read file", "code": "read_failed"}, status=500
        )
    _, f, size, content_type, path = result

    try:
        start, end = 0, size - 1
        status = 200
        range_header = request.headers.get("Range")
        if range_header:
            parsed = _parse_range_header(range_header, size)
            if parsed is None:
                _log("denied", path)
                return web.json_response(
                    {"error": "range not satisfiable", "code": "bad_range"},
                    status=416,
                    headers={"Content-Range": f"bytes */{size}"},
                )
            start, end = parsed
            status = 206

        resp = web.StreamResponse(status=status)
        resp.content_type = content_type
        resp.content_length = end - start + 1
        resp.headers["Accept-Ranges"] = "bytes"
        resp.headers["X-Content-Type-Options"] = "nosniff"
        if status == 206:
            resp.headers["Content-Range"] = f"bytes {start}-{end}/{size}"
        # SEL: record the ALLOW decision before any bytes move. prepare() and
        # the write loop can be cancelled by a client disconnect, and a
        # permitted read must never leave the audit trail empty because the
        # client hung up first.
        _log("success", path)
        await resp.prepare(request)

        await asyncio.to_thread(f.seek, start)
        remaining = end - start + 1
        while remaining > 0:
            try:
                chunk = await asyncio.to_thread(f.read, min(_STREAM_CHUNK_BYTES, remaining))
            except OSError:
                # A mid-stream filesystem error must leave a SEL outcome; the
                # response is already streaming so all we can do is stop short.
                _log("failure", path)
                raise
            if not chunk:
                break  # file truncated under us; the announced length just ends short
            remaining -= len(chunk)
            try:
                await resp.write(chunk)
            except (ConnectionResetError, ConnectionError):
                break  # client hung up (scrubbing, tab close) -- normal for media
        with contextlib.suppress(Exception):
            await resp.write_eof()
        return resp
    finally:
        # close() can wait on the buffered-file lock while a worker-thread
        # read is in flight (task cancellation), so it must not run on the
        # event loop either.
        with contextlib.suppress(Exception):
            await asyncio.to_thread(f.close)


def _file_write_blocking(path: str, content: str) -> str | None:
    """Replace *path*'s contents atomically, carrying its access controls.

    Returns ``None`` on success or ``"notfound"`` when the target was rejected;
    any other failure propagates for the caller to log.

    Split out of :func:`api_file_write` so the whole transaction runs OFF the
    event loop. Every call in here is a blocking filesystem call, and on a
    network-backed path (an SMB share, a stalled FUSE mount) each one can take
    seconds, which on the loop thread freezes chat and the heartbeat alongside
    it. Being on a worker thread also re-arms the Windows rename retry inside
    ``atomic_write``, which deliberately degrades to a single attempt when it
    finds a running loop in its own thread.

    Routing through ``open_access_control_source`` rather than a bare ``os.open``
    is what keeps this working on Windows: it returns ``None`` where the xattr
    syscalls do not exist, and a read handle held open across the write would
    make ``os.replace`` fail with ``PermissionError`` on every save there.

    ``path`` is already canonicalized by ``_validate_dashboard_path``
    (``realpath``), so its final component is symlink-free and the helper's
    ``O_NOFOLLOW`` rejects nothing legitimate -- it closes the window where that
    component is swapped for a link after the check. That refusal is a rejected
    target rather than a server fault, hence ``"notfound"`` and not an exception.
    """
    # Pin the parent chain FIRST, then address the leaf only through that
    # descriptor. The pin is what stops atomic_write's temp create and publishing
    # rename from re-resolving the parent by name, and the ORDER is what stops the
    # metadata read below from re-resolving it either: a directory replaced at
    # that name between the pin and the leaf open would otherwise supply the mode
    # and ACL while the write published into the pinned original.
    #
    # pin_parent, NOT open_dir_pinned: ``path`` is already realpath-canonicalized,
    # so every component of its parent was a real directory at validation time.
    # pin_parent walks THAT recorded chain with O_NOFOLLOW per component, so a
    # component swapped for a link since is REFUSED. open_dir_pinned would
    # realpath the chain again here and follow the swap instead -- a fresh
    # resolution cannot be more faithful than the one already done, only less.
    #
    # None on a platform that cannot walk a parent by descriptor or cannot stage
    # and rename through one, where atomic_write keeps the by-name floor. Both
    # probes are asked because they are two capabilities: atomic_write refuses a
    # descriptor it cannot use rather than silently writing by name.
    dir_fd: int | None = None
    if pinned_fs.supports_pinned_walk() and pinned_parent_replace_supported():
        try:
            dir_fd = pinned_fs.pin_parent(os.path.dirname(path), what="file directory")
        except (pinned_fs.PinnedPathRefusal, OSError):
            # Both are the same disposition -- a target that can no longer be
            # reached through the tree the caller validated is rejected, not a
            # server fault -- so they share one arm rather than drifting apart.
            return "notfound"
    src_fd: int | None = None
    try:
        try:
            src_fd = open_access_control_source(path, dir_fd=dir_fd)
        except OSError:
            return "notfound"
        # os.stat by name only where nothing was pinned: with dir_fd the helper
        # always hands back a descriptor, so the mode comes from the same inode
        # the ACL does and neither is re-resolved.
        src_stat = os.fstat(src_fd) if src_fd is not None else os.stat(path)
        # mode= keeps the previous copymode behaviour (permission bits), and
        # preserve_access_control_from is ADDITIVE to it: copymode carried BITS
        # only, so a named POSIX ACL (system.posix_acl_access) the owner set was
        # silently dropped the moment the replace installed a fresh inode. The
        # carry is allowlisted to the ACL and user.* names -- it must NOT replay a
        # privilege-bearing security.capability onto caller-supplied content.
        atomic_write(
            path,
            content,
            mode=_stat_mod.S_IMODE(src_stat.st_mode),
            preserve_access_control_from=src_fd,
            parent_dir_fd=dir_fd,
        )
    finally:
        for fd in (src_fd, dir_fd):
            if fd is not None:
                try:
                    os.close(fd)
                except OSError:
                    pass
    return None


async def api_file_write(request: web.Request) -> web.Response:
    """POST /api/file-write — write file content from the markdown panel."""
    from kiro_crew.validation import (  # noqa: F811
        FILE_WRITE_SCHEMA,
        ValidationError,
        validate_tool_args,
    )

    # max_bytes=None: the body carries the file's whole contents, which has no
    # defensible byte ceiling (issue #5587 sweep).
    body, body_err = await read_bounded_json(request, max_bytes=None)
    if body_err is not None:
        return body_err
    assert body is not None  # read_bounded_json returns (dict, None) on success

    try:
        validate_tool_args(
            {"path": body.get("path", ""), "content": body.get("content", "")}, FILE_WRITE_SCHEMA
        )
    except ValidationError:
        _sel().log_tool_invocation(
            session_key="dashboard",
            tool_name="file_write",
            outcome="denied",
            resources=body.get("path", ""),
        )
        return web.json_response({"error": "invalid input"}, status=400)

    path = _validate_dashboard_path(body.get("path", ""))
    if not path:
        _sel().log_tool_invocation(
            session_key="dashboard",
            tool_name="file_write",
            outcome="denied",
            resources=body.get("path", ""),
        )
        return web.json_response({"error": "invalid or forbidden path"}, status=400)
    if not os.path.isfile(path):
        _sel().log_tool_invocation(
            session_key="dashboard", tool_name="file_write", outcome="not_found", resources=path
        )
        return web.json_response({"error": "not found"}, status=404)
    try:
        # Off the event loop: see _file_write_blocking's own note on why the
        # whole transaction is offloaded rather than each call individually.
        outcome = await asyncio.to_thread(_file_write_blocking, path, body.get("content", ""))
        if outcome == "notfound":
            _sel().log_tool_invocation(
                session_key="dashboard",
                tool_name="file_write",
                outcome="not_found",
                resources=path,
            )
            return web.json_response({"error": "not found", "code": "not_found"}, status=404)
        _sel().log_tool_invocation(
            session_key="dashboard", tool_name="file_write", outcome="success", resources=path
        )
        return web.json_response({"ok": True})
    except Exception:
        logging.getLogger(__name__).exception("file_write failed for %s", path)
        _sel().log_tool_invocation(
            session_key="dashboard", tool_name="file_write", outcome="failure", resources=path
        )
        return web.json_response({"error": "failed to write file"}, status=500)


def _subsequence_run(q: str, haystack: str) -> tuple[int, int]:
    """Greedily match ``q`` as a subsequence of ``haystack``.

    Returns how many of ``q``'s characters were consumed in order, and the
    longest run of matches that landed on consecutive ``haystack`` positions
    within that single greedy pass -- NOT the longest contiguous occurrence of
    ``q``, since the scan never backtracks over an earlier isolated match
    (``q="ab"`` against ``"axxab"`` consumes both chars but reports a run of
    1). A consumed count below ``len(q)`` means ``haystack`` does not contain
    ``q`` as a subsequence at all; the caller normalizes the run length by
    ``len(q)`` into the contiguity term of the fuzzy score.
    """
    qi = 0
    consecutive = 0
    max_run = 0
    for ch in haystack:
        if qi < len(q) and ch == q[qi]:
            qi += 1
            consecutive += 1
            max_run = max(max_run, consecutive)
        else:
            consecutive = 0
    return qi, max_run


def _fuzzy_score(q: str, name: str, rel: str) -> float:
    """Score a file match. Higher = better. Returns 0 for no match."""
    nl = name.lower()
    rl = rel.lower()
    score = 0.0

    # Exact filename match (sans extension)
    stem = nl.rsplit(".", 1)[0] if "." in nl else nl
    if q == nl or q == stem:
        score += 100.0
    elif nl.startswith(q):
        score += 50.0
    elif q in nl:
        score += 30.0
    elif q in rl:
        score += 10.0
    else:
        # Fuzzy: check whether the query chars appear in order in the
        # filename, falling back to the search-root-relative path when the
        # filename alone does not carry the query as an in-order subsequence.
        matched_on_name = True
        qi, max_run = _subsequence_run(q, nl)
        if qi < len(q):
            matched_on_name = False
            qi, max_run = _subsequence_run(q, rl)
        if qi < len(q):
            return 0.0  # not all query chars found
        # Score based on coverage ratio and longest consecutive run
        matched_len = len(nl) if matched_on_name else len(rl)
        coverage = len(q) / max(matched_len, 1)
        score += 5.0 + 15.0 * (max_run / len(q)) + 5.0 * coverage

    # Bonus: shorter filenames are more relevant
    score += max(0.0, 5.0 - len(nl) * 0.1)
    return score


async def api_file_search(request: web.Request) -> web.Response:
    """GET /api/file-search?q=... — fuzzy filename search for the @-mention file picker."""
    # Re-imported at call time (not reused from the module-level binding) so a
    # test that stubs ``kiro_crew.security.is_sensitive_path`` is observed by the
    # project-root rejection below.
    from kiro_crew.security import is_sensitive_path  # noqa: F811

    caller = request.get("user", "dashboard")
    query = request.query.get("q", "").strip().lower()
    if len(query) < 2:
        return web.json_response({"results": []})

    # Result page size. Default mirrors SEARCH_RESULT_CAP in FolderPanel.tsx;
    # the caller may raise it via ``limit`` (the folder panel's expand control),
    # clamped to ``_SEARCH_LIMIT_CEILING`` server-side. Non-integer input falls
    # back to the default, mirroring how ``kinds`` handles unknown values.
    try:
        max_results = int(request.query.get("limit", "15"))
    except ValueError:
        max_results = 15
    max_results = max(1, min(max_results, _SEARCH_LIMIT_CEILING))

    # kinds: "all" (default) returns both files and directories; "files" or
    # "dirs" restricts the result set. Unknown values fall back to "all".
    kinds = request.query.get("kinds", "all").strip().lower()
    if kinds not in ("all", "files", "dirs"):
        kinds = "all"
    want_files = kinds in ("all", "files")
    want_dirs = kinds in ("all", "dirs")

    # Scope search to project (arbitrary path) or workspace
    project = request.query.get("project", "")
    ws_name = request.query.get("workspace", "")
    search_roots: list[str] = []
    if project:
        project = os.path.realpath(os.path.expanduser(project))
        if is_sensitive_path(project):
            _sel().log_api_access(caller=caller, operation="file_search", outcome="denied", resources=project, error="sensitive path")
            return web.json_response({"error": "Access denied"}, status=403)
        if os.path.isdir(project):
            search_roots.append(project)
        else:
            return web.json_response(
                {"results": [], "error": "Project directory not found"}, status=404
            )
    elif ws_name:
        from kiro_crew.config.loader import workspace_dir_for  # noqa: F811
        ws_path = str(workspace_dir_for(ws_name))
        if os.path.isdir(ws_path):
            search_roots.append(ws_path)

    scoped = bool(search_roots)

    if not search_roots:
        # Fallback: project dir, then the kirocrew workspace.
        #
        # Bare $HOME is deliberately NOT a fallback root. Walking it reaches
        # every TCC-gated folder macOS knows about, and each one costs a
        # separate consent dialog -- paid on an unscoped keystroke the user
        # never pointed anywhere. The results did not justify it either: the
        # walk stops at max_scan entries in os.walk order, so an unscoped home
        # search returned whichever files happened to be reached first rather
        # than the best matches. Callers that genuinely want home can still
        # ask for it explicitly with ?project=$HOME, which is scoped and
        # searched in full.
        proj = os.environ.get("KIROCREW_PROJECT_DIR", "")
        if proj and os.path.isdir(proj):
            search_roots.append(proj)
        mc_workspace = str(data_home() / "workspace")
        if os.path.isdir(mc_workspace):
            search_roots.append(mc_workspace)

    # Filter out sensitive roots
    safe_roots: list[str] = []
    for r in search_roots:
        if is_sensitive_path(r):
            _sel().log_api_access(caller=caller, operation="file_search", outcome="denied", resources=r, error="sensitive path")
        else:
            safe_roots.append(r)

    # Fast path: use in-memory index when available for a single scoped project
    state: DashboardState = request.app["state"]
    if scoped and len(safe_roots) == 1:
        idx = state.file_indexes.get(safe_roots[0])
        if idx and idx.is_ready and not idx.truncated:
            results = await asyncio.to_thread(idx.search, query, _fuzzy_score, max_results, kinds)
            trimmed = [{k: v for k, v in r.items() if k != "_score"} for r in results]
            _sel().log_api_access(caller=caller, operation="file_search", outcome="allowed", resources=f"q={query} kinds={kinds} indexed=true entries={idx.entry_count} results={len(trimmed)}")
            return web.json_response({"results": trimmed, "root": safe_roots[0]})

    # Fallback: walk filesystem per request
    # Dot-prefixed FILES stay excluded (startswith(".") guard in _collect).
    # Dot-prefixed DIRECTORIES (.github, .kiro, .claude) ARE offered as
    # candidates; only skip_dirs below are dropped from both descent and results.
    # skip_dirs is the SAME shared set the indexed fast path uses (imported from
    # file_index), so the two paths of this endpoint cannot diverge on which
    # directories are suppressed -- see #5677.
    skip_dirs = _WALK_SKIP_DIRS

    max_scan = _WALK_MAX_SCAN_SCOPED if scoped else _WALK_MAX_SCAN_UNSCOPED
    max_collect = max_results * 10  # collect enough candidates for good scoring, then stop

    def _walk_file_search() -> list[dict]:
        """Blocking file-system walk — offloaded via asyncio.to_thread.

        Files and directories are collected into SEPARATE candidate lists, each
        with its own ``max_collect`` allowance. A shared list would let a burst
        of matching directories fill the cap before the files in the same
        directory are even examined, dropping the likely target before the
        file-before-dir tie-break ever runs. Files are also scanned first at each
        level, so under a tight scan budget the file candidates are the ones that
        survive.

        An independent ``_WALK_MAX_DIRS_VISITED`` ceiling bounds how many
        directories the walk descends into, so no request can traverse a whole
        large tree.
        """
        found: dict[str, list[dict]] = {"file": [], "dir": []}
        walked: dict[str, int] = {"file": 0, "dir": 0}
        dirs_visited = 0
        wanted = {"file": want_files, "dir": want_dirs}

        def _done(kind: str) -> bool:
            return (
                not wanted[kind]
                or walked[kind] >= max_scan
                or len(found[kind]) >= max_collect
            )

        def _full() -> bool:
            return dirs_visited >= _WALK_MAX_DIRS_VISITED or (_done("file") and _done("dir"))

        def _collect(kind: str, dirpath: str, names: list[str], root_dir: str) -> None:
            """Score and collect one kind of entry from a single directory level."""
            for name in names:
                if _done(kind):
                    return
                walked[kind] += 1
                if kind == "file" and name.startswith("."):
                    continue
                full = os.path.join(dirpath, name)
                score = _fuzzy_score(query, name, os.path.relpath(full, root_dir))
                if score <= 0:
                    continue
                # Resolve symlinks before the sensitivity check so a link into a
                # sensitive tree cannot slip through.
                if is_sensitive_path(os.path.realpath(full)):
                    continue
                try:
                    st = os.stat(full)
                except OSError:
                    continue
                found[kind].append({
                    "path": full,
                    "name": name,
                    "kind": kind,
                    "size": st.st_size if kind == "file" else 0,
                    "mtime": int(st.st_mtime),
                    "_score": score,
                })

        for root_dir in safe_roots:
            if _full():
                break
            # macOS: prune the TCC-gated folders. Reaching into them would pop
            # one consent modal PER folder. ``scoped`` means the user NAMED
            # this root (?project= / ?workspace=), so even ``project=$HOME``
            # is deliberate and is searched in full.
            for dirpath, dirnames, filenames in os.walk(root_dir):
                # Bounds the traversal; the per-kind counters stop advancing once
                # their kind is done.
                dirs_visited += 1
                # A dot-prefixed directory (.github, .kiro, .claude) should be
                # OFFERED as a candidate even though we must not DESCEND into it.
                # These were previously conflated: ``dirnames`` was pruned in
                # place (dropping dot-dirs) before _collect saw it.
                #
                # Build the candidate list (offered AND stat'd) first, then
                # derive the narrower descent list from it. Both drop skip_dirs
                # (.git, node_modules, ...). On an UNSCOPED root the TCC-gated
                # folders (Downloads, Desktop, Library, ... from a $HOME root on
                # macOS) must also be dropped from candidates -- merely offering
                # one means os.stat-ing it, which pops a consent modal; a scoped
                # root is deliberate and is never TCC-pruned, matching the
                # descent rule below. Only the leading-dot rule differs: a dot-
                # dir is a valid candidate but is removed from the descent list.
                base_dirs = [d for d in dirnames if d not in skip_dirs]
                if scoped:
                    candidate_dirs = base_dirs
                else:
                    candidate_dirs = platform_compat.tcc_prune_walk_dirs(
                        root_dir, dirpath, base_dirs
                    )
                dirnames[:] = [d for d in candidate_dirs if not d.startswith(".")]
                # Files first: under a tight scan budget the file candidates are
                # the ones that survive.
                _collect("file", dirpath, filenames, root_dir)
                _collect("dir", dirpath, candidate_dirs, root_dir)
                if _full():
                    break
        return found["file"] + found["dir"]

    results = await asyncio.to_thread(_walk_file_search)

    # Sort by score descending, files before dirs on a tie, then shorter name, then recency
    now = time.time()
    results.sort(key=lambda r: (
        -r["_score"], r["kind"] == "dir", len(r["name"]), now - r["mtime"],
    ))

    # Strip internal scoring field before response
    trimmed = [{k: v for k, v in r.items() if k != "_score"} for r in results[:max_results]]

    _sel().log_api_access(caller=caller, operation="file_search", outcome="allowed", resources=f"q={query} kinds={kinds} roots={len(safe_roots)} results={len(trimmed)}")
    return web.json_response({
        "results": trimmed,
        "root": safe_roots[0] if scoped and safe_roots else "",
    })


async def api_file_diff(request: web.Request) -> web.Response:
    """GET /api/file-diff?path=... — returns git diff and HEAD content for a file."""
    raw_path = request.query.get("path", "").strip()
    if not raw_path:
        _sel().log_api_access(caller=request.get("user", "dashboard"), operation="file_diff", outcome="allowed", resources="empty_path")
        return web.json_response({"diff": "", "original": ""})
    raw_path = os.path.realpath(os.path.expanduser(raw_path))
    if not os.path.isfile(raw_path):
        _sel().log_api_access(caller=request.get("user", "dashboard"), operation="file_diff", outcome="allowed", resources=f"path={raw_path}", error="not_found")
        return web.json_response({"diff": "", "original": ""})
    if is_sensitive_path(raw_path):
        _sel().log_api_access(caller=request.get("user", "dashboard"), operation="file_diff", outcome="denied", resources=raw_path, error="sensitive path")
        return web.json_response({"error": "Access denied"}, status=403)

    dirpath = os.path.dirname(raw_path)

    def _run() -> dict:
        # Disable textconv/filter drivers and fsmonitor to prevent code execution
        # via .gitattributes or .git/config in untrusted repos.
        _git = ["git", "-c", "diff.textconv=", "-c", "core.attributesFile=/dev/null", "-c", "core.fsmonitor="]
        _env = {**os.environ, "GIT_ATTR_NOSYSTEM": "1"}
        try:
            subprocess.run(
                [*_git, "rev-parse", "--git-dir"],
                cwd=dirpath, capture_output=True, text=True, encoding="utf-8",
                errors="replace", timeout=5, check=True, env=_env,
            )
        except (subprocess.TimeoutExpired, subprocess.CalledProcessError, FileNotFoundError, UnicodeDecodeError):
            # Only a failed repository preflight may claim "not a git repo":
            # the client renders not_git as "there is no baseline", which is a
            # statement about the file, not about git's health. Failures past
            # this point (a timeout on a slow repo, git disappearing mid-flight)
            # are computation failures and must report "error" instead.
            return {"diff": "", "original": "", "status": "not_git"}
        try:
            # Get HEAD content
            root = subprocess.run(
                [*_git, "rev-parse", "--show-toplevel"],
                cwd=dirpath, capture_output=True, text=True, encoding="utf-8",
                errors="replace", timeout=5, env=_env,
            ).stdout.strip()
            rel = os.path.relpath(raw_path, root)
            head = subprocess.run(
                [*_git, "show", "--no-textconv", f"HEAD:{rel}"],
                cwd=dirpath, capture_output=True, text=True, encoding="utf-8",
                errors="replace", timeout=10, env=_env,
            )
            original = head.stdout if head.returncode == 0 else ""
            # Get diff
            r = subprocess.run(
                [*_git, "diff", "--no-textconv", "--no-ext-diff", "HEAD", "--", raw_path],
                cwd=dirpath, capture_output=True, text=True, encoding="utf-8",
                errors="replace", timeout=10, env=_env,
            )
            diff = r.stdout.strip() if r.returncode == 0 else ""
            if not diff:
                # Check for untracked file
                r2 = subprocess.run(
                    [*_git, "status", "--porcelain", "--", raw_path],
                    cwd=dirpath, capture_output=True, text=True, encoding="utf-8",
                    errors="replace", timeout=5, env=_env,
                )
                if r2.returncode == 0 and r2.stdout.strip().startswith("??"):
                    r3 = subprocess.run(
                        [*_git, "diff", "--no-textconv", "--no-ext-diff", "--no-index", "/dev/null", raw_path],
                        cwd=dirpath, capture_output=True, text=True, encoding="utf-8",
                        errors="replace", timeout=10, env=_env,
                    )
                    diff = r3.stdout if r3.stdout else ""
                    return {"diff": diff, "original": "", "status": "untracked"}
            if r.returncode != 0:
                # `git diff` failed and the untracked probe above did not claim
                # the file. This must stay distinguishable from a genuinely
                # unmodified file: falling through would report status "clean",
                # presenting a git failure as "no changes" — a false negative on
                # a question users act on. The probe runs FIRST because the
                # dominant non-zero exit is `fatal: bad revision 'HEAD'` in a
                # freshly-initialized repo with no commits, where every file is
                # simply untracked and the all-added diff is the true answer.
                # Still HTTP 200: the request succeeded, only the diff did not.
                return {"diff": "", "original": original, "status": "error"}
            status = "modified" if diff else "clean"
            return {"diff": diff, "original": original, "status": status}
        except (subprocess.TimeoutExpired, subprocess.CalledProcessError, FileNotFoundError, UnicodeDecodeError):
            return {"diff": "", "original": "", "status": "error"}

    result = await asyncio.to_thread(_run)
    _sel().log_api_access(caller=request.get("user", "dashboard"), operation="file_diff", outcome="allowed", resources=f"path={raw_path}")
    return web.json_response(result)


def _browse_dirs_sync(base: str, skip: set[str]) -> list[dict]:
    """Walk *base* one level deep and return its visible subdirectories.

    Blocking, and unboundedly so: *base* is caller-chosen and defaults to ``$HOME``,
    so the scan is as large as that directory, and every surviving entry additionally
    pays an ``is_sensitive_path`` call that resolves several paths of its own. Run via
    ``asyncio.to_thread`` so one large directory cannot hold the sole event loop for
    the duration of the listing.
    """
    dirs: list[dict] = []
    try:
        for entry in sorted(os.scandir(base), key=lambda e: e.name.lower()):
            if entry.is_dir(follow_symlinks=True) and entry.name not in skip and not entry.name.startswith("."):
                # Resolve symlinks before the sensitivity check — a symlink in
                # a benign dir pointing at ~/.aws would otherwise pass through.
                if is_sensitive_path(os.path.realpath(entry.path)):
                    continue
                dirs.append({"name": entry.name, "path": entry.path})
    except PermissionError:
        pass
    return dirs


def _browse_files_sync(base: str, skip: set[str]) -> tuple[list[dict], list[dict]]:
    """Walk *base* one level deep and return its ``(dirs, files)`` entries.

    The sibling of :func:`_browse_dirs_sync` and blocking for the same reasons, plus a
    ``stat`` per entry for the mtime the browser sorts on. Offloaded the same way.
    """
    dirs: list[dict] = []
    files: list[dict] = []
    try:
        # Sort: dirs before files, then alphabetical
        for entry in sorted(os.scandir(base), key=lambda e: (not e.is_dir(follow_symlinks=True), e.name.lower())):
            if entry.name.startswith("."):
                continue
            # Resolve symlinks before the sensitivity check — a symlink in a
            # benign dir pointing at ~/.aws would otherwise pass through.
            if is_sensitive_path(os.path.realpath(entry.path)):
                continue
            # Capture mtime so the activity-panel browser can offer a
            # sort-by-date option; fall back to 0 on a race (entry removed
            # mid-scan) so one unstattable entry never breaks the listing.
            try:
                mtime = int(entry.stat(follow_symlinks=True).st_mtime)
            except OSError:
                mtime = 0
            if entry.is_dir(follow_symlinks=True):
                if entry.name not in skip:
                    dirs.append({"name": entry.name, "path": entry.path, "mtime": mtime})
            elif entry.is_file(follow_symlinks=True):
                files.append({"name": entry.name, "path": entry.path, "mtime": mtime})
    except PermissionError:
        pass
    return dirs, files


async def api_browse_dirs(request: web.Request) -> web.Response:
    """GET /api/browse-dirs?path=... — list subdirectories for directory browser."""
    caller = request.get("user", "dashboard")
    raw = request.query.get("path", "").strip()
    base = os.path.realpath(os.path.expanduser(raw)) if raw else os.path.realpath(os.path.expanduser("~"))
    if not os.path.isdir(base):
        return web.json_response({"error": "Not a directory", "path": base}, status=400)
    if is_sensitive_path(base):
        _sel().log_api_access(caller=caller, operation="browse_dirs", outcome="denied", resources=base, error="sensitive path")
        return web.json_response({"error": "Access denied"}, status=403)
    skip = {".git", "node_modules", "__pycache__", ".cache", ".venv", "venv", "env", ".kirocrew", ".kiro", ".aim"}
    dirs = await asyncio.to_thread(_browse_dirs_sync, base, skip)
    _sel().log_api_access(caller=caller, operation="browse_dirs", outcome="allowed", resources=base)
    return web.json_response({"path": base, "parent": os.path.dirname(base), "dirs": dirs})


#: Depth ceiling for the walk-up that looks for a repository root. A project
#: directory nested deeper than this below its repo root is reported as
#: not-a-repo rather than paying an unbounded number of stat calls per request.
_GIT_ROOT_WALK_LIMIT = 40

#: A HEAD file is one short line; cap the read so a hostile symlink to something
#: enormous cannot be slurped into memory.
_HEAD_READ_LIMIT = 4096


def _read_git_meta_prefix(path: str) -> str | None:
    """Read a bounded prefix of a git metadata file through the hooks gate.

    ``.git`` and ``.git/HEAD`` are ordinary filesystem paths inside a directory
    the caller chose, so either can be a symlink pointing at something the
    gateway must never read — a secret whose first line happens to look like a
    ref, or a 40-64 char hex blob that would match the detached-HEAD shape.
    ``hooks.safe_read_prefix`` canonicalises via realpath, refuses sensitive
    resolved targets, and opens with ``O_NOFOLLOW`` as TOCTOU defence against a
    final-component swap. A refused or unreadable path returns ``None`` and the
    caller degrades to "no branch".
    """
    data = safe_read_prefix(path, _HEAD_READ_LIMIT)
    if data is None:
        return None
    return data.decode("utf-8", errors="replace").strip()


def _git_head_path(root: str) -> str | None:
    """Resolve the HEAD file for the repo at *root*.

    A linked worktree's ``.git`` is a FILE containing ``gitdir: <path>``, and that
    directory holds the worktree's own HEAD — so the pointer has to be followed
    rather than assuming ``<root>/.git`` is a directory.
    """
    dot = os.path.join(root, ".git")
    if os.path.isdir(dot):
        return os.path.join(dot, "HEAD")
    pointer = _read_git_meta_prefix(dot)
    if pointer is None or not pointer.startswith("gitdir:"):
        return None
    gitdir = pointer.split(":", 1)[1].strip()
    if not gitdir:
        return None
    if not os.path.isabs(gitdir):
        gitdir = os.path.join(root, gitdir)
    return os.path.join(gitdir, "HEAD")


def _slot_project_snapshot(state: DashboardState) -> list[str]:
    """Copy every live slot's project dir. MUST run on the event loop.

    Slots are created and deleted by other coroutines on the loop, so the copy
    has to happen where those mutations are serialised against it. Doing it in a
    worker thread would iterate a dict that the loop can mutate underneath.
    Pure in-memory, no I/O — safe to call inline.
    """
    dirs: list[str] = []
    for slot in list(getattr(state, "_slots", {}).values()):
        proj = getattr(slot, "project", "") or ""
        if proj:
            dirs.append(proj)
    return dirs


def _known_project_dirs(slot_projects: list[str]) -> list[str]:
    """Server-held project directories a branch lookup may be asked about.

    The caller's slot snapshot plus the recorded recent-projects list —
    directories the gateway itself set or the user already picked through the
    project picker. Nothing in the returned list comes from the current request.
    Reads a file, so this belongs in a worker thread.
    """
    dirs: list[str] = list(slot_projects)
    fp = config_dir() / "recent_projects.json"
    try:
        recent = json.loads(fp.read_text(encoding="utf-8")) if fp.is_file() else []
    except (OSError, ValueError):
        recent = []
    if isinstance(recent, list):
        dirs.extend(d for d in recent if isinstance(d, str) and d)
    return dirs


def _match_known_project(raw: str, known: list[str]) -> str | None:
    """Map a request-supplied path onto the matching known project directory.

    Returns the SERVER-HELD string, never the caller's, so request data is only
    ever a comparison operand and never reaches a filesystem call. Matching is
    pure string normalisation (expanduser + normpath) with no filesystem access
    on the untrusted value — deliberately not realpath, which would stat a
    caller-controlled path and reintroduce the probe this guard removes.
    """
    want = os.path.normpath(os.path.expanduser(raw))
    for cand in known:
        if os.path.normpath(os.path.expanduser(cand)) == want:
            return cand
    return None


def _project_git_branch(base: str) -> dict:
    """Resolve the checked-out branch for ``base``.

    Returns ``{"repo": False}`` when ``base`` is not inside a git repository.
    For a repository, returns the repo root plus either a ``branch`` name or,
    on a detached HEAD, ``detached: True`` with the short commit in ``head``.
    """
    root: str | None = None
    cur = base
    for _ in range(_GIT_ROOT_WALK_LIMIT):
        # A worktree's .git is a FILE (a gitdir pointer), not a directory, so
        # probe for existence rather than is_dir() — otherwise every KiroCrew
        # worktree reports as not-a-repo.
        if os.path.exists(os.path.join(cur, ".git")):
            root = cur
            break
        parent = os.path.dirname(cur)
        if parent == cur:
            break
        cur = parent
    if root is None:
        return {"repo": False}
    # ``root`` is derived from an allow-listed project directory, but a directory
    # NAME is itself agent-influenceable via set_project and this value is echoed
    # to the dashboard, so it goes through the same egress redaction as the branch
    # label. A normal path is unchanged.
    out: dict = {"repo": True, "repoRoot": redact(root)}
    head_path = _git_head_path(root)
    if head_path is None:
        return out
    raw = _read_git_meta_prefix(head_path)
    if raw is None:
        # Unreadable, absent, or refused by the sensitive-path gate: still a
        # repo, just no label.
        return out
    if raw.startswith("ref:"):
        ref = raw[len("ref:"):].strip()
        prefix = "refs/heads/"
        if ref.startswith(prefix) and len(ref) > len(prefix):
            # Branch names are attacker/agent-controllable content that this route
            # renders in the dashboard AND makes copyable, so it goes through the
            # canonical egress redaction like any other echoed string. Ordinary
            # branch names are unchanged; one that embeds something matching a
            # credential pattern is masked rather than displayed.
            out["branch"] = redact(ref[len(prefix):])
        return out
    # A bare object id in HEAD means detached (mid-rebase, bisect, explicit
    # --detach). Surface a short form so the caller shows something truthful
    # instead of an empty label. This is a fixed 7-char prefix rather than git's
    # dynamic uniqueness-based abbreviation — for a decorative label that is an
    # acceptable difference, and it needs no repository query.
    if re.fullmatch(r"[0-9a-fA-F]{40,64}", raw):
        out["detached"] = True
        out["head"] = redact(raw[:7])
    return out


def _match_known_project_for(slot_projects: list[str], raw: str) -> str | None:
    """Build the allow-list and match *raw* against it. Worker-thread only.

    Takes an already-taken slot snapshot rather than the live state, so nothing
    here touches structures the event loop mutates. Both remaining halves must
    stay off the loop: reading the recent-projects file does I/O, and
    ``expanduser`` on a ``~user`` form does a passwd lookup, which can block on
    NSS/LDAP for an authenticated caller passing ``?path=~x/y``.
    """
    return _match_known_project(raw, _known_project_dirs(slot_projects))


def _resolve_project_git(project: str) -> tuple[str, str, dict]:
    """Vet *project* and read its branch. Runs entirely in a worker thread.

    Every filesystem touch for the request lives here: ``realpath``,
    the directory check, and ``is_sensitive_path`` all stat, so a project on a
    stalled network mount would block the event loop for the whole probe if any
    of them ran inline.

    Returns ``(status, base, info)`` with status ``"ok"``, ``"not_a_dir"``, or
    ``"sensitive"``; ``info`` is populated only for ``"ok"``.
    """
    base = os.path.realpath(os.path.expanduser(project))
    if not os.path.isdir(base):
        return "not_a_dir", base, {}
    if is_sensitive_path(base):
        return "sensitive", base, {}
    return "ok", base, _project_git_branch(base)


async def api_project_git(request: web.Request) -> web.Response:
    """GET /api/project/git?path=... — checked-out branch for a project dir.

    ``path`` is matched against the gateway's own set of known project
    directories and the matched server-held value is what gets stat'd, so this
    route cannot be used to probe arbitrary filesystem paths for existence or
    git metadata. An unrecognised directory is refused outright.
    """
    state: DashboardState = request.app["state"]
    caller = request.get("user", "dashboard")
    raw = request.query.get("path", "").strip()
    if not raw:
        return web.json_response({"error": "path required"}, status=400)
    project = await asyncio.to_thread(
        _match_known_project_for, _slot_project_snapshot(state), raw
    )
    if project is None:
        _sel().log_api_access(
            caller=caller,
            operation="project_git",
            outcome="denied",
            resources=raw,
            error="not a known project directory",
        )
        return web.json_response({"error": "Unknown project directory"}, status=403)
    status, base, info = await asyncio.to_thread(_resolve_project_git, project)
    if status == "not_a_dir":
        # Redacted like every other echoed path: this arm is reachable whenever a
        # known project directory is deleted or replaced between the allow-list
        # match and the stat, so it is a live egress surface, not a dead branch.
        return web.json_response(
            {"error": "Not a directory", "path": redact(base)}, status=400
        )
    if status == "sensitive":
        _sel().log_api_access(
            caller=caller,
            operation="project_git",
            outcome="denied",
            resources=base,
            error="sensitive path",
        )
        return web.json_response({"error": "Access denied"}, status=403)
    _sel().log_api_access(
        caller=caller, operation="project_git", outcome="allowed", resources=base
    )
    # The SEL audit above records the real path; the response body is an egress
    # surface the dashboard renders, so the echoed path is redacted like the rest.
    return web.json_response({"path": redact(base), **info})


async def api_browse_files(request: web.Request) -> web.Response:
    """GET /api/browse-files?path=... — list files and subdirectories for the activity-panel file browser.

    Mirrors api_browse_dirs security model (sensitive-path filtering, access logging,
    skip set for build artifacts) but returns files alongside directories. Entries
    are sorted dirs-first then alphabetically; hidden files and common build dirs
    are skipped.
    """
    caller = request.get("user", "dashboard")
    raw = request.query.get("path", "").strip()
    base = os.path.realpath(os.path.expanduser(raw)) if raw else os.path.realpath(os.path.expanduser("~"))
    if not os.path.isdir(base):
        return web.json_response({"error": "Not a directory", "path": base}, status=400)
    if is_sensitive_path(base):
        _sel().log_api_access(caller=caller, operation="browse_files", outcome="denied", resources=base, error="sensitive path")
        return web.json_response({"error": "Access denied"}, status=403)
    skip = {".git", "node_modules", "__pycache__", ".cache", ".venv", "venv", "env", ".kirocrew", ".kiro", ".aim", "build", "dist", ".next"}
    dirs, files = await asyncio.to_thread(_browse_files_sync, base, skip)
    _sel().log_api_access(caller=caller, operation="browse_files", outcome="allowed", resources=base)
    return web.json_response({"path": base, "parent": os.path.dirname(base), "dirs": dirs, "files": files})


async def api_dashboard_config(request: web.Request) -> web.Response:
    """GET/PUT /api/dashboard/config — read or write dashboard settings."""
    from kiro_crew.config.loader import KiroCrewConfig  # noqa: F811

    # Owner gate for PUT: reject non-owner writes before paying the config-load
    # I/O cost. The check is cheap (in-memory predicate + optional off-thread
    # SEL audit on denial) compared to the KiroCrewConfig.load() thread hop
    # below, so non-owner PUT requests are rejected immediately.
    if request.method == "PUT":
        from kiro_crew.dashboard.handlers._shared import require_owner_dashboard_request

        owner_denied = await require_owner_dashboard_request(request, "dashboard_config.write")
        if owner_denied is not None:
            return owner_denied

    # Offloaded: KiroCrewConfig.load() stats, reads, parses, and validates config
    # files. The client polls this endpoint on an interval to pick up externally
    # edited dashboard.gitlab_hosts, so a slow or network-backed config directory
    # would otherwise stall the sole event loop on every poll.
    try:
        cfg = await asyncio.to_thread(KiroCrewConfig.load)
    except asyncio.CancelledError:
        # A cancellation at this await (client disconnect mid-poll, gateway
        # shutdown) would otherwise unwind the handler before either the
        # read-success or the write-success/failure audit below, leaving an
        # authorized config access attempt entirely absent from the
        # tamper-evident SEL chain. Pair the landed request with an explicit
        # failure event, then re-raise so cancellation still propagates.
        _sel().log_tool_invocation(
            session_key="dashboard",
            tool_name=(
                "dashboard_config_write" if request.method == "PUT" else "dashboard_config_read"
            ),
            outcome="failure",
            error="request_cancelled",
        )
        raise
    if request.method == "PUT":
        # Default cap: the body is a fixed set of dashboard toggles and numbers.
        body, body_err = await read_bounded_json(request)
        if body_err is not None:
            _sel().log_tool_invocation(
                session_key="dashboard",
                tool_name="dashboard_config_write",
                outcome="failure",
                error=_body_err_code(body_err),
            )
            return body_err
        assert body is not None  # read_bounded_json returns (dict, None) on success
        _allowed = {"restore_sessions", "restore_window_minutes", "merge_queued_messages", "widget_density", "use_builtin_browser", "verbosity", "quick_send", "session_grid", "tail_fork_enabled", "link_previews", "mcp_app_panel", "auto_open_git_panel", "folder_suggestions_enabled", "session_card_source_links"}
        # One-release backward-compat shim for removed key; delete after all clients update.
        deprecated_ignored_keys = {"tail_fork_head_handling"}
        # Read-only keys the GET exposes: both settings surfaces save with
        # `mutate({ ...dashCfg, ...patch })`, so every GET field comes back in the
        # PUT body. Drop them here instead of listing them in _allowed -- they
        # stay unwritable, but a round-tripped read-only field must not 400 an
        # unrelated toggle save.
        read_only_ignored_keys = {"gitlab_hosts", "jira_hosts", "social_share_enabled"}
        body = {
            k: v
            for k, v in body.items()
            if k not in deprecated_ignored_keys and k not in read_only_ignored_keys
        }
        unknown = set(body.keys()) - _allowed
        if unknown:
            _sel().log_tool_invocation(
                session_key="dashboard", tool_name="dashboard_config_write", outcome="failure"
            )
            return web.json_response({"error": f"Unknown fields: {unknown}"}, status=400)
        updates: dict[str, object] = {}
        if "restore_sessions" in body:
            val = body["restore_sessions"]
            if not isinstance(val, bool):
                _sel().log_tool_invocation(
                    session_key="dashboard", tool_name="dashboard_config_write", outcome="failure"
                )
                return web.json_response(
                    {"error": "restore_sessions must be a boolean"}, status=400
                )
            updates["restore_sessions"] = val
        try:
            if "restore_window_minutes" in body:
                updates["restore_window_minutes"] = max(
                    0, min(1440, int(body["restore_window_minutes"]))
                )
        except (TypeError, ValueError):
            _sel().log_tool_invocation(
                session_key="dashboard", tool_name="dashboard_config_write", outcome="failure"
            )
            return web.json_response(
                {"error": "restore_window_minutes must be an integer"}, status=400
            )
        if "merge_queued_messages" in body:
            val = body["merge_queued_messages"]
            if not isinstance(val, bool):
                _sel().log_tool_invocation(
                    session_key="dashboard", tool_name="dashboard_config_write", outcome="failure"
                )
                return web.json_response(
                    {"error": "merge_queued_messages must be a boolean"}, status=400
                )
            updates["merge_queued_messages"] = val
        if "widget_density" in body:
            val = body["widget_density"]
            if val not in ("more", "less"):
                _sel().log_tool_invocation(
                    session_key="dashboard", tool_name="dashboard_config_write", outcome="failure"
                )
                return web.json_response(
                    {"error": "widget_density must be 'more' or 'less'"}, status=400
                )
            updates["widget_density"] = val
        # Apply ONLY when it is the sole submitted setting. The Browser panel
        # sends it alone; the Chat settings panel PUTs the whole config object
        # from its own (possibly stale) cache, and applying it on that path would
        # let a Chat-panel save silently revert a toggle another client changed
        # (lost update).
        if body.keys() == {"use_builtin_browser"}:
            val = body["use_builtin_browser"]
            if not isinstance(val, bool):
                _sel().log_tool_invocation(
                    session_key="dashboard", tool_name="dashboard_config_write", outcome="failure"
                )
                return web.json_response(
                    {
                        "error": "use_builtin_browser must be a boolean",
                        "code": "invalid_use_builtin_browser",
                    },
                    status=400,
                )
            updates["use_builtin_browser"] = val
        if "verbosity" in body:
            val = body["verbosity"]
            if val not in ("default", "concise", "ultra", "answer_only"):
                _sel().log_tool_invocation(
                    session_key="dashboard", tool_name="dashboard_config_write", outcome="failure"
                )
                return web.json_response(
                    {
                        "error": (
                            "verbosity must be 'default', 'concise', 'ultra' "
                            "or 'answer_only'"
                        )
                    },
                    status=400,
                )
            updates["verbosity"] = val
        if "tail_fork_enabled" in body:
            val = body["tail_fork_enabled"]
            if not isinstance(val, bool):
                _sel().log_tool_invocation(
                    session_key="dashboard", tool_name="dashboard_config_write", outcome="failure"
                )
                return web.json_response(
                    {"error": "tail_fork_enabled must be a boolean"}, status=400
                )
            updates["tail_fork_enabled"] = val
        if "folder_suggestions_enabled" in body:
            val = body["folder_suggestions_enabled"]
            if not isinstance(val, bool):
                _sel().log_tool_invocation(
                    session_key="dashboard", tool_name="dashboard_config_write", outcome="failure"
                )
                return web.json_response(
                    {
                        "error": "folder_suggestions_enabled must be a boolean",
                        "code": "invalid_folder_suggestions_enabled",
                    },
                    status=400,
                )
            updates["folder_suggestions_enabled"] = val
        if "link_previews" in body:
            val = body["link_previews"]
            if not isinstance(val, bool):
                _sel().log_tool_invocation(
                    session_key="dashboard", tool_name="dashboard_config_write", outcome="failure"
                )
                return web.json_response(
                    {
                        "error": "link_previews must be a boolean",
                        "code": "invalid_link_previews",
                    },
                    status=400,
                )
            updates["link_previews"] = val
        if "quick_send" in body:
            val = body["quick_send"]
            if not isinstance(val, bool):
                _sel().log_tool_invocation(
                    session_key="dashboard", tool_name="dashboard_config_write", outcome="failure"
                )
                return web.json_response(
                    {"error": "quick_send must be a boolean"}, status=400
                )
            updates["quick_send"] = val
        if "session_grid" in body:
            val = body["session_grid"]
            if not isinstance(val, bool):
                _sel().log_tool_invocation(
                    session_key="dashboard", tool_name="dashboard_config_write", outcome="failure"
                )
                return web.json_response(
                    {"error": "session_grid must be a boolean"}, status=400
                )
            updates["session_grid"] = val
        if "mcp_app_panel" in body:
            val = body["mcp_app_panel"]
            if not isinstance(val, bool):
                _sel().log_tool_invocation(
                    session_key="dashboard", tool_name="dashboard_config_write", outcome="failure"
                )
                return web.json_response(
                    {
                        "error": "mcp_app_panel must be a boolean",
                        "code": "invalid_mcp_app_panel",
                    },
                    status=400,
                )
            updates["mcp_app_panel"] = val
        if "auto_open_git_panel" in body:
            val = body["auto_open_git_panel"]
            if not isinstance(val, bool):
                _sel().log_tool_invocation(
                    session_key="dashboard", tool_name="dashboard_config_write", outcome="failure"
                )
                return web.json_response(
                    {
                        "error": "auto_open_git_panel must be a boolean",
                        "code": "invalid_auto_open_git_panel",
                    },
                    status=400,
                )
            updates["auto_open_git_panel"] = val
        if "session_card_source_links" in body:
            val = body["session_card_source_links"]
            if not isinstance(val, bool):
                _sel().log_tool_invocation(
                    session_key="dashboard", tool_name="dashboard_config_write", outcome="failure"
                )
                return web.json_response(
                    {
                        "error": "session_card_source_links must be a boolean",
                        "code": "invalid_session_card_source_links",
                    },
                    status=400,
                )
            updates["session_card_source_links"] = val
        # Serialize the read-modify-write under BOTH config locks so no concurrent
        # writer -- in-process OR another process -- can clobber it:
        #  * update_config_locked holds the cross-process advisory file lock
        #    (<config>.lock) for the whole read-modify-write, so a concurrent
        #    `kirocrew config set` (which takes that same file lock) cannot land
        #    between our read and write and be silently discarded.
        #  * wrapping it in _get_config_lock() (the repo-wide, loop-bound asyncio
        #    lock) serializes it against the legacy in-process writers that still
        #    save under that asyncio lock alone.
        # Both run OFF-THREAD so the event loop is never blocked. Only the
        # dashboard.<field> keys this request validated are written, leaving every
        # other config section on disk untouched. GET stays lock-free.
        from kiro_crew.config.loader import update_config_locked  # noqa: F811
        from kiro_crew.dashboard.handlers.agents import (  # lazy: import cycle
            _get_config_lock,
        )

        def _apply_dashboard_updates(data: dict) -> dict:
            # `dashboard` is normally a dict; tolerate a missing or malformed
            # (non-dict, e.g. a hand-edited/corrupt `[]`) section by replacing it
            # with a fresh dict rather than raising TypeError mid-write. The prior
            # non-dict value carried no valid dashboard settings, so this recovers
            # the section instead of losing data, and leaves other config keys
            # untouched.
            section = data.get("dashboard")
            if not isinstance(section, dict):
                section = data["dashboard"] = {}
            for _field, _value in updates.items():
                section[_field] = _value
            return data

        try:
            async with _get_config_lock():
                await asyncio.to_thread(
                    lambda: update_config_locked(mutate=_apply_dashboard_updates)
                )
        except asyncio.CancelledError:
            # Cancellation (client disconnect / gateway shutdown) during the
            # off-thread write does NOT hit the `except Exception` below
            # (CancelledError is a BaseException), and the worker may still land
            # the write -- so the authorized attempt would vanish from the SEL
            # chain. Log a failure outcome, then re-raise so cancellation still
            # propagates. Mirrors the load guard above; both satisfy the
            # backend-security-controls audit contract.
            _sel().log_tool_invocation(
                session_key="dashboard",
                tool_name="dashboard_config_write",
                outcome="failure",
                error="request_cancelled",
            )
            raise
        except Exception:
            # Any other failure to land the write -- e.g. a corrupt on-disk config
            # makes update_config_locked's fail-closed read raise ConfigReadError
            # (not an OSError, so nothing else catches it) -- must still leave a
            # tamper-evident SEL entry rather than escaping as an unlogged 500.
            _sel().log_tool_invocation(
                session_key="dashboard", tool_name="dashboard_config_write", outcome="failure"
            )
            logger.exception("dashboard config write failed")
            return web.json_response(
                {
                    "error": "failed to save dashboard config",
                    "code": "dashboard_config_write_failed",
                },
                status=500,
            )
        _sel().log_tool_invocation(
            session_key="dashboard", tool_name="dashboard_config_write", outcome="success"
        )
        chips_written = updates.get("session_card_source_links")
        if isinstance(chips_written, bool):
            # Publish the new value NOW instead of leaving it to the next
            # allowlist refresh. That refresh is on a 30s TTL, so without this
            # the sidebar keeps rendering chips for up to half a minute after an
            # explicit click -- the switch acknowledges itself instantly and
            # nothing appears to happen, which reads as broken. This handler
            # already knows the value, so polling for it is the wrong shape.
            #
            # The push is the other half: the publisher bumps the shared
            # generation, but the owner websocket only compares that generation
            # once per TTL round, so a push here is what re-serializes the slots
            # with the new answer.
            #
            # The value is read OUTSIDE the try on purpose: only the publish and
            # the push may fail silently, so a body that never carried this key
            # cannot reach the publisher at all -- and a test can tell the two
            # apart instead of a swallowed KeyError standing in for the guard.
            try:
                from kiro_crew.dashboard.handlers.source_providers import (  # lazy: import cycle
                    publish_session_card_chips_now,
                )

                await publish_session_card_chips_now(chips_written)
                state = request.app.get("state")
                if state is not None:
                    state.push_slots_update()
            except Exception:
                # Best-effort: the write itself succeeded, and the next refresh
                # round picks the value up within one TTL. Failing the request
                # here would report a saved setting as unsaved.
                logger.debug("chip-switch snapshot publish failed", exc_info=True)
        return web.json_response({"ok": True})
    _sel().log_tool_invocation(
        session_key="dashboard", tool_name="dashboard_config_read", outcome="success"
    )
    # Governance-derived, not a config value: the dashboard draws the "Share as
    # image" entry only when this is true, and it has no other way to know — the
    # share card has no server-side action to refuse, so this read IS the
    # enforcement point. Resolved off-thread (profile resolution may read from
    # disk); every decision is SEL-audited by the probe itself.
    from kiro_crew.dashboard import social_share

    social_share_denied = await asyncio.to_thread(social_share.is_share_denied)
    return web.json_response(
        {
            "restore_sessions": cfg.dashboard.restore_sessions,
            "restore_window_minutes": cfg.dashboard.restore_window_minutes,
            "merge_queued_messages": cfg.dashboard.merge_queued_messages,
            "widget_density": cfg.dashboard.widget_density,
            "use_builtin_browser": cfg.dashboard.use_builtin_browser,
            "verbosity": cfg.dashboard.verbosity,
            "quick_send": cfg.dashboard.quick_send,
            "session_grid": cfg.dashboard.session_grid,
            "mcp_app_panel": cfg.dashboard.mcp_app_panel,
            "auto_open_git_panel": cfg.dashboard.auto_open_git_panel,
            "session_card_source_links": cfg.dashboard.session_card_source_links,
            "tail_fork_enabled": cfg.dashboard.tail_fork_enabled,
            "link_previews": cfg.dashboard.link_previews,
            "folder_suggestions_enabled": cfg.dashboard.folder_suggestions_enabled,
            # Read-only here (absent from the PUT allowlist above): authorizing a
            # self-managed GitLab instance is a config-file decision, not a
            # dashboard toggle. The client uses it only to decide which pasted
            # links become source tabs; the provider handler re-checks every URL.
            "gitlab_hosts": list(cfg.dashboard.gitlab_hosts),
            # Same discipline for Jira: Atlassian Cloud (*.atlassian.net) is
            # auto-recognized; self-hosted instances need explicit allowlisting.
            "jira_hosts": list(cfg.dashboard.jira_hosts),
            # Read-only: the `capabilities.social_share` governance answer. False
            # withdraws the "Share as image" menu entry; there is no toggle behind
            # it, so nothing here is writable.
            "social_share_enabled": not social_share_denied,
        }
    )


# ── /api/file-sheet: xlsx → JSON cell grid ───────────────────────────────────
# Caps bound what one request can materialize server-side and ship to the
# browser. 500 rows matches CsvViewer's display cap so the two table viewers
# truncate consistently. The member/expansion caps bound zip inflation: the
# on-disk size cap only limits the COMPRESSED archive, and a crafted workbook
# can expand orders of magnitude larger than it stores.
_SHEET_MAX_SHEETS = 20
_SHEET_MAX_ROWS = 500
_SHEET_MAX_COLS = 100
_SHEET_MAX_MEMBERS = 4096
_SHEET_MAX_EXPANDED_BYTES = 200 * 1024 * 1024
# Text amplification caps. Shared strings are stored once in the archive but
# referenced per cell, so the expansion cap above does not bound the RESPONSE:
# one 32 KiB string referenced by every cell would amplify into gigabytes of
# JSON. Cells truncate individually, and the whole workbook gets a cumulative
# text budget past which the preview refuses (the frontend degrades to the
# download card).
_SHEET_MAX_CELL_CHARS = 2000
_SHEET_MAX_TEXT_CHARS = 5 * 1000 * 1000


class _SheetRefusal(Exception):
    """Deliberate refusal carrying its HTTP status and machine-readable code;
    raised on the worker thread and mapped to a response by api_file_sheet."""

    def __init__(self, status: int, message: str, code: str):
        super().__init__(message)
        self.status = status
        self.code = code


def _sheet_cell_json(value: object) -> object:
    """Serialize one workbook cell value into a JSON-safe primitive."""
    if value is None or isinstance(value, (bool, int)):
        return value
    if isinstance(value, str):
        # Workbook text is file content leaving the host through the dashboard
        # — same egress class as api_file_read, so the same redaction applies.
        # Redact BEFORE truncating so the scan always sees the complete text,
        # then cap the cell so one shared string cannot bloat every row.
        value = redact(value)
        if len(value) > _SHEET_MAX_CELL_CHARS:
            return value[:_SHEET_MAX_CELL_CHARS] + "…"
        return value
    if isinstance(value, float):
        # NaN/Infinity are rejected by JSON.parse in the browser; the stdlib
        # encoder would happily emit the JS-only tokens.
        if value != value or value in (float("inf"), float("-inf")):
            return str(value)
        return value
    if isinstance(value, _dt.datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, (_dt.date, _dt.time)):
        return value.isoformat()
    return redact(str(value))


def _sheet_formula_text(value: object) -> str | None:
    """Return the formula source ("=…") for a formula-pass cell value, else None."""
    text: object = value
    if not (isinstance(text, str) and text.startswith("=")):
        # Array formulas come back as openpyxl ArrayFormula objects carrying .text.
        text = getattr(value, "text", None)
    if isinstance(text, str) and text.startswith("="):
        text = redact(text)
        if len(text) > _SHEET_MAX_CELL_CHARS:
            return text[:_SHEET_MAX_CELL_CHARS] + "…"
        return text
    return None


def _load_sheet_payload(f: BinaryIO, *, max_bytes: int) -> dict:
    """Read, vet, and parse the workbook into the sheet-grid payload.

    Runs ENTIRELY on a worker thread (via asyncio.to_thread) so filesystem
    latency, the first (heavy) openpyxl import, and parse time never stall the
    gateway event loop. Receives the checked-open file object from
    :func:`_open_checked_file` (the shared open-and-check prefix, which owns
    path validation, the sensitive-path gate, and the symlink-refusing open)
    and takes ownership: the file is closed on every path. The bounded-read
    cap is this endpoint's size policy, passed in as *max_bytes*. openpyxl is
    a soft import: absence surfaces as ImportError from this thread and the
    handler maps it to 501.
    """
    import io

    with f:
        import openpyxl  # noqa: F401  (probe here, off-loop; parse imports lazily too)

        # Bounded read is the size guard: a pre-check via fstat would race a
        # concurrent writer (the file can grow between the stat and the read,
        # e.g. an agent still generating the workbook), while reading at most
        # cap+1 bytes bounds memory unconditionally.
        data = f.read(max_bytes + 1)
    if len(data) > max_bytes:
        raise _SheetRefusal(413, "file too large", "file_too_large")
    header = data[:4]
    # OOXML spreadsheets are ZIP containers; refuse anything else before
    # openpyxl touches the bytes.
    if not header.startswith(b"PK\x03\x04"):
        raise _SheetRefusal(415, "not an OOXML spreadsheet", "not_a_spreadsheet")
    # Vet the archive's declared inventory before anything inflates it --
    # including ZipFile construction itself, which materializes one ZipInfo
    # per central-directory entry. The EOCD preflight bounds that allocation
    # from the raw bytes; the infolist() pass then bounds what openpyxl can
    # actually expand (zipfile truncates each member at its declared
    # file_size, so the central directory's numbers are authoritative).
    _vet_zip_eocd(data)
    with zipfile.ZipFile(io.BytesIO(data)) as zf:
        infos = zf.infolist()
        if (
            len(infos) > _SHEET_MAX_MEMBERS
            or sum(i.file_size for i in infos) > _SHEET_MAX_EXPANDED_BYTES
        ):
            raise _SheetRefusal(413, "workbook expands too large", "workbook_expands_too_large")
    return _parse_workbook_grid(data)


# Generous per-entry allowance for the central-directory size preflight: a
# record is 46 bytes plus name/extra/comment, and OOXML part names are short.
_SHEET_MAX_CDIR_ENTRY_BYTES = 512


def _vet_zip_eocd(data: bytes) -> None:
    """Refuse archives whose end-of-central-directory record declares an
    oversized inventory, BEFORE zipfile.ZipFile is constructed.

    Delegates to the shared vet (kiro_crew.zip_vet) so this endpoint, knowledge
    ingest, and document parsing share one implementation of the preflight --
    only the caps and the error channel stay per-caller. This endpoint's
    observable behaviour is unchanged: a tail with no usable EOCD still reads as
    "not a spreadsheet" (415), an over-cap inventory as an expansion refusal
    (413).
    """
    try:
        vet_zip_inventory_bytes(
            data,
            max_members=_SHEET_MAX_MEMBERS,
            max_cdir_entry_bytes=_SHEET_MAX_CDIR_ENTRY_BYTES,
        )
    except ZipInventoryRejected as exc:
        if exc.reason in ("missing_eocd", "truncated_eocd", "unreadable"):
            raise _SheetRefusal(
                415, "not an OOXML spreadsheet", "not_a_spreadsheet") from exc
        raise _SheetRefusal(
            413, "workbook expands too large", "workbook_expands_too_large") from exc


def _parse_workbook_grid(data: bytes) -> dict:
    """Parse xlsx bytes into a JSON-safe sheet grid. Runs on a worker thread.

    The workbook is loaded twice in read-only streaming mode: once with
    data_only=True (formula cells yield the value cached by the writing
    application) and once with data_only=False (formula cells yield the
    formula source). Cells prefer the cached value; when a file carries no
    cache — typical for openpyxl-generated workbooks — the formula text is
    shown instead of an empty cell. Both loads stream the same bytes, so the
    row structures are identical and can be zipped in lockstep.
    """
    import io
    import itertools

    from openpyxl import load_workbook

    wb_vals = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    wb_form = load_workbook(io.BytesIO(data), read_only=True, data_only=False)
    try:
        names = wb_vals.sheetnames
        sheets: list[dict] = []
        # Cumulative post-truncation text budget across the whole workbook:
        # shared strings are stored once but referenced per cell, so archive
        # size caps alone do not bound the JSON response this grid becomes.
        text_chars = 0
        for name in names[:_SHEET_MAX_SHEETS]:
            ws_v, ws_f = wb_vals[name], wb_form[name]
            if not hasattr(ws_v, "iter_rows"):  # chartsheets have no cell grid
                continue
            # Dimension records can lie (some writers emit a stale ref such as
            # A1:A1 for a populated sheet); read-only mode trusts them, so
            # iter_rows would stop early and silently truncate the preview.
            # Force a real scan of each sheet instead.
            if hasattr(ws_v, "reset_dimensions"):
                ws_v.reset_dimensions()
                ws_f.reset_dimensions()
            raw: list[list[object]] = []
            rows_truncated = False
            cols_truncated = False
            paired = zip(ws_v.iter_rows(values_only=True), ws_f.iter_rows(values_only=True))
            for vrow, frow in itertools.islice(paired, _SHEET_MAX_ROWS + 1):
                if len(raw) >= _SHEET_MAX_ROWS:
                    rows_truncated = True
                    # No total is reported: any count derived from workbook
                    # geometry is attacker-influenced (a single sparse row at
                    # index 1e9 makes the read-only reader synthesize a
                    # billion empties), so nothing here iterates past the cap.
                    break
                if len(vrow) > _SHEET_MAX_COLS:
                    cols_truncated = True
                out: list[object] = []
                for vv, fv in list(zip(vrow, frow))[:_SHEET_MAX_COLS]:
                    ftxt = _sheet_formula_text(fv)
                    cell = ftxt if (vv is None and ftxt) else _sheet_cell_json(vv)
                    if isinstance(cell, str):
                        text_chars += len(cell)
                        if text_chars > _SHEET_MAX_TEXT_CHARS:
                            raise _SheetRefusal(
                                413, "workbook text too large to preview",
                                "workbook_text_too_large",
                            )
                    out.append(cell)
                raw.append(out)
            # Trim trailing all-empty rows, then normalize every row to the
            # widest non-empty extent so the client renders a rectangle.
            while raw and all(c is None or c == "" for c in raw[-1]):
                raw.pop()
            width = 0
            for r in raw:
                w = len(r)
                while w and (r[w - 1] is None or r[w - 1] == ""):
                    w -= 1
                width = max(width, w)
            rows = [r[:width] + [None] * (width - len(r[:width])) for r in raw] if width else []
            sheets.append({
                # Names take the same redact+truncate path as cell text — a
                # crafted workbook.xml can carry arbitrarily long sheet names.
                "name": _sheet_cell_json(name),
                "rows": rows,
                "truncated_rows": rows_truncated,
                "truncated_cols": cols_truncated,
            })
        return {
            "sheets": sheets,
            "total_sheets": len(names),
            "truncated_sheets": len(names) > _SHEET_MAX_SHEETS,
        }
    finally:
        wb_vals.close()
        wb_form.close()


async def api_file_sheet(request: web.Request) -> web.Response:
    """GET /api/file-sheet?path=… — parse an OOXML spreadsheet into a JSON cell grid.

    Powers the file viewer's inline xlsx preview. The security prefix is the
    shared :func:`_open_checked_file` (dashboard path validation,
    sensitive-path block, a symlink-refusing open — _open_rb_nofollow: atomic
    O_NOFOLLOW on POSIX, lstat guard on Windows); this endpoint's own policy
    on top is the bounded-read size cap, the zip-expansion caps, and a ZIP
    magic-byte check before openpyxl touches the bytes. All file IO and
    parsing runs on a worker thread so a large workbook cannot stall the
    event loop, and cell text is credential-redacted like every other
    dashboard egress. openpyxl is soft-imported: without it the endpoint
    answers 501 and the frontend degrades to the download card.
    """

    def _log(outcome: str, res: str) -> None:
        _sel().log_tool_invocation(
            session_key="dashboard", tool_name="file_sheet", outcome=outcome, resources=res,
        )

    raw_path = request.query.get("path", "")
    # The validated path once the prefix produces one -- exported by the
    # worker callback so the exception handlers log the same SEL resource
    # the success path does.
    res_path = raw_path

    def _open_and_load() -> dict | _OpenDenied:
        """Open-and-check plus parse, in ONE worker-thread hop.

        The checked open file object never crosses back to the event loop:
        every path that opens it also closes it on THIS thread (refusals
        close inside the prefix; the parser's ``with f:`` covers the rest).
        A cancellation of the awaiting task therefore cannot strand an open
        file in a discarded future or finalize one on the loop -- the
        future's result is only ever a payload dict or a typed refusal.
        """
        nonlocal res_path
        checked = _open_checked_file(
            raw_path, tool_name="file_sheet", log_open_failure=False,
        )
        if isinstance(checked, _OpenDenied):
            return checked
        res_path = checked.path
        return _load_sheet_payload(checked.file, max_bytes=_MAX_UPLOAD_BYTES)

    try:
        result = await asyncio.to_thread(_open_and_load)
    except asyncio.CancelledError:
        # Shutdown or client disconnect: the access attempt must not vanish
        # from the audit trail. No resource handling here -- the worker
        # callback owns the file's whole lifetime.
        _log("cancelled", res_path)
        raise
    except ImportError:
        # openpyxl absent: the preview is unavailable, not broken. The probe
        # runs inside the worker thread so even the first heavy import never
        # touches the event loop.
        _log("failure", res_path)
        return web.json_response(
            {"error": "spreadsheet preview unavailable", "code": "preview_unavailable"},
            status=501,
        )
    except _SheetRefusal as refusal:
        # Both refusal kinds map to literal statuses so the response shape
        # stays statically checkable; the carried code names the exact cause.
        _log("denied", res_path)
        if refusal.status == 415:
            return web.json_response({"error": str(refusal), "code": refusal.code}, status=415)
        return web.json_response({"error": str(refusal), "code": refusal.code}, status=413)
    except OSError:
        # Read failure on the already-checked fd. (A symlink never reaches
        # here: the shared prefix refuses it as _OpenDenied("symlink_refused")
        # before the parser sees a file object.)
        _log("failure", res_path)
        return web.json_response({"error": "cannot read file", "code": "read_failed"}, status=500)
    except Exception:
        # openpyxl's failure surface is wide (bad zip members, malformed XML,
        # unexpected workbook parts). Every parse failure degrades to the same
        # client answer, and the frontend falls back to the download card.
        logger.warning("file-sheet: cannot parse workbook %s", res_path, exc_info=True)
        _log("failure", res_path)
        return web.json_response(
            {"error": "cannot parse workbook", "code": "parse_failed"}, status=422
        )
    if isinstance(result, _OpenDenied):
        code, res = result.code, result.path
        if code == "invalid_path":
            _log("denied", res)
            return web.json_response(
                {"error": "invalid or forbidden path", "code": "invalid_path"}, status=400
            )
        if code == "sensitive_path":
            _log("denied", res)
            return web.json_response(
                {"error": "sensitive path blocked", "code": "sensitive_path"}, status=403
            )
        if code == "not_found":
            _log("not_found", res)
            return web.json_response({"error": "not found", "code": "not_found"}, status=404)
        if code == "symlink_refused":
            _log("denied", res)
            return web.json_response(
                {"error": "symlinks not allowed", "code": "symlink_refused"}, status=403
            )
        if code == "file_too_large":
            # Reachable only if this endpoint ever passes fstat_cap; mapped so
            # a policy refusal can never masquerade as the 500 below. (Its
            # size guard today is the bounded read inside _load_sheet_payload.)
            _log("denied", res)
            return web.json_response(
                {"error": "file too large", "code": "file_too_large"}, status=413
            )
        # read_failed: the residual code.
        _log("failure", res)
        return web.json_response({"error": "cannot read file", "code": "read_failed"}, status=500)
    _log("success", res_path)
    return web.json_response(result)


# ── Git status & log endpoints ──────────────────────────────────────────────


# Ceiling on captured git stdout for the Git-panel endpoints. Status output is
# repo-content-sized (an agent-authored repo can make it arbitrarily large) and
# these endpoints are POLLED by the dashboard, so an unbounded
# ``capture_output=True`` buffer is a memory-DoS surface. 8 MB comfortably
# holds the 500-file slice the responses return while bounding the worst case.
_GIT_PANEL_STDOUT_CAP = 8 * 1024 * 1024


def _run_git_bounded(
    args: list[str], cwd: str, env: dict, timeout: float,
    cap: int = _GIT_PANEL_STDOUT_CAP,
) -> tuple[int, str, bool]:
    """Run git capturing at most ``cap`` bytes of stdout.

    Returns ``(returncode, stdout_text, truncated)``. When the process
    outlives ``timeout`` or overflows ``cap`` it is killed and reported as
    truncated with a nonzero returncode -- callers already treat nonzero as
    "no data", which is the safe degraded answer for a pathological repo.
    """
    # OS-sandbox + credential-scrubbed env chokepoint (worktree.py's _run_git
    # pattern): the repository content is agent-influenced, and git filter
    # drivers (filter.<name>.clean/process from .git/config) can run during
    # status re-hashing -- ``-c`` flags cannot neutralize arbitrary driver
    # names, so isolation, not argument hygiene, is the containment. Fail
    # CLOSED: no sandbox backend means no data, not an unisolated spawn.
    cleanup: str | None = None
    try:
        argv, env, cleanup = sandboxed_spawn_argv(args, mode="strict", env=env)
    except RuntimeError:
        return -9, "", False
    env["GIT_TERMINAL_PROMPT"] = "0"
    try:
        try:
            proc = popen_limited(
                argv, cwd=cwd, env=env,
                stdout=subprocess.PIPE, stderr=subprocess.DEVNULL,
            )
        except OSError:
            # The cwd (project dir) can vanish between the handler's isdir
            # check and this spawn, and the git binary itself can be absent.
            # Both are "no data", never a 500 out of a polling endpoint.
            return -9, "", False
        buf = bytearray()
        overflow = False

        def _drain() -> None:
            nonlocal overflow
            assert proc.stdout is not None
            while True:
                chunk = proc.stdout.read(65536)
                if not chunk:
                    return
                if len(buf) + len(chunk) > cap:
                    buf.extend(chunk[: cap - len(buf)])
                    overflow = True
                    return
                buf.extend(chunk)

        reader = threading.Thread(target=_drain, daemon=True)
        reader.start()
        reader.join(timeout)
        timed_out = reader.is_alive()
        if timed_out or overflow:
            proc.kill()
            reader.join(5)
        try:
            rc = proc.wait(timeout=5)
        except subprocess.TimeoutExpired:
            proc.kill()
            rc = -9
        if timed_out or overflow:
            rc = rc or -9
        return rc, bytes(buf).decode("utf-8", "replace"), timed_out or overflow
    finally:
        if cleanup:
            with contextlib.suppress(OSError):
                os.unlink(cleanup)


def _porcelain_unquote(path: str) -> str:
    """Decode a C-quoted porcelain v1 path (``"foo \\"bar\\""`` -> ``foo "bar"``).

    Porcelain v1 wraps a path in double quotes and backslash-escapes it when it
    contains quotes, backslashes, or control characters (``core.quotePath=false``
    already keeps plain non-ASCII raw). Returning the quoted display form would
    point the row -- and a subsequent open/save -- at a file that does not
    exist. Decode failures fall back to the raw string rather than raising.
    """
    if len(path) < 2 or not (path.startswith('"') and path.endswith('"')):
        return path
    body = path[1:-1]
    out = bytearray()
    i = 0
    escapes = {"n": 10, "t": 9, "r": 13, "a": 7, "b": 8, "f": 12, "v": 11,
               "\\": 92, '"': 34}
    while i < len(body):
        ch = body[i]
        if ch != "\\":
            out.extend(ch.encode("utf-8"))
            i += 1
            continue
        if i + 1 >= len(body):
            return path  # dangling escape: not valid quoting, keep raw
        nxt = body[i + 1]
        if nxt in escapes:
            out.append(escapes[nxt])
            i += 2
        elif nxt.isdigit() and i + 3 < len(body) + 1 and body[i + 1:i + 4].isdigit():
            out.append(int(body[i + 1:i + 4], 8) & 0xFF)
            i += 4
        else:
            return path
    return out.decode("utf-8", "replace")


# Repo-scoped config keys that hand git a program to run when it touches file
# content (status re-hashes modified files through ``filter.<name>.clean``).
# ``-c`` cannot neutralize arbitrary driver names, so a repo declaring one is
# refused outright — the same fail-closed stance as worktree.py's
# ``_checkout_filter``.
_GIT_FILTER_KEY_RE = re.compile(
    r"^filter\..+\.(process|smudge|clean)$", re.IGNORECASE
)


def _repo_declares_filter_driver(git_cmd: list[str], base: str, env: dict) -> bool:
    """True when repo-supplied config names a content-filter driver (or the
    probe cannot prove it does not).

    Mirrors ``worktree.py::_checkout_filter``: drivers can only come from a
    config file the repository supplies — ``--local`` (``.git/config``) and,
    when ``extensions.worktreeConfig`` is on, ``--worktree``
    (``$GIT_DIR/config.worktree``). ``--includes`` is mandatory: a specific-scope
    query defaults include-following OFF, so a driver reached through
    ``include.path`` would be invisible to the probe yet still execute.
    Global/system config is deliberately not probed (the user's own machine
    setup, e.g. ``git lfs install``, is not repository-supplied). A probe that
    fails refuses: an unreadable scope cannot be proven filter-free. The probe
    itself is safe — ``git config`` reads files and never runs drivers.
    """
    scopes = ["--local"]
    ext_rc, ext_out, _ = _run_git_bounded(
        [*git_cmd, "config", "--bool", "--get", "extensions.worktreeConfig"],
        cwd=base, env=env, timeout=5,
    )
    if ext_rc == 0 and ext_out.strip() == "true":
        scopes.append("--worktree")
    for scope in scopes:
        rc, out, _ = _run_git_bounded(
            [*git_cmd, "config", scope, "--includes", "--name-only", "--list"],
            cwd=base, env=env, timeout=5,
        )
        if rc != 0:
            return True
        for key in out.splitlines():
            if _GIT_FILTER_KEY_RE.match(key.strip()):
                return True
    return False


async def api_project_git_status(request: web.Request) -> web.Response:
    """GET /api/project/git/status?path=... - working tree status for a project dir.

    Returns staged/unstaged/untracked files with per-file line-change counts.
    Path must match a known project directory (same allow-list as api_project_git).
    """
    state: DashboardState = request.app["state"]
    caller = request.get("user", "dashboard")
    raw = request.query.get("path", "").strip()
    if not raw:
        return web.json_response({"error": "path required", "code": "path_required"}, status=400)
    project = await asyncio.to_thread(
        _match_known_project_for, _slot_project_snapshot(state), raw
    )
    if project is None:
        _sel().log_api_access(
            caller=caller,
            operation="project_git_status",
            outcome="denied",
            resources=raw,
            error="not a known project directory",
        )
        return web.json_response({"error": "Unknown project directory", "code": "unknown_project_dir"}, status=403)

    base = await asyncio.to_thread(
        lambda: os.path.realpath(os.path.expanduser(project))
    )
    # Both probes stat the filesystem (a stalled network mount would block the
    # event loop), so they run in a worker thread like the realpath above.
    if await asyncio.to_thread(is_sensitive_path, base):
        _sel().log_api_access(
            caller=caller,
            operation="project_git_status",
            outcome="denied",
            resources=base,
            error="sensitive path",
        )
        return web.json_response({"error": "Access denied", "code": "access_denied"}, status=403)
    # Log the allow decision here (not after _run) so every authorized access
    # is audited, including the not-a-directory / not-a-repo early answers.
    _sel().log_api_access(
        caller=caller, operation="project_git_status", outcome="allowed", resources=base
    )
    if not await asyncio.to_thread(os.path.isdir, base):
        return web.json_response({"repo": False, "files": []})

    def _run() -> dict:
        _git_cmd = [
            "git",
            "-c", "diff.textconv=",
            "-c", "core.attributesFile=/dev/null",
            "-c", f"core.hooksPath={os.devnull}",
            "-c", "core.fsmonitor=",
            # Repo-local .gitattributes is still consulted despite the
            # attributesFile override, so keep driver escape hatches shut and
            # emit non-ASCII paths raw (UTF-8) instead of C-quoted so the
            # panel can open them.
            "-c", "core.quotePath=false",
        ]
        _env = {**os.environ, "GIT_ATTR_NOSYSTEM": "1"}

        # Check if it's a repo
        probe_rc, _probe_out, _ = _run_git_bounded(
            [*_git_cmd, "rev-parse", "--git-dir"], cwd=base, env=_env, timeout=5,
        )
        if probe_rc != 0:
            return {"repo": False, "files": []}

        # Refuse repos whose own config names a content-filter driver: status
        # re-hashes modified files through ``filter.<name>.clean``, which would
        # execute that program on every 5s poll. Degraded-but-safe empty answer.
        if _repo_declares_filter_driver(_git_cmd, base, _env):
            return {"repo": True, "files": []}

        # Get repo root and branch info
        root_rc, root_out, _ = _run_git_bounded(
            [*_git_cmd, "rev-parse", "--show-toplevel"], cwd=base, env=_env, timeout=5,
        )
        repo_root = root_out.strip() if root_rc == 0 else base

        # Branch + ahead/behind via status -b
        status_rc, status_out, _ = _run_git_bounded(
            [*_git_cmd, "status", "--porcelain=v1", "-b", "--untracked-files=all"],
            cwd=base, env=_env, timeout=10,
        )
        if status_rc != 0:
            return {"repo": True, "repoRoot": repo_root, "files": []}

        lines = status_out.splitlines()
        branch = None
        ahead = 0
        behind = 0

        # Parse the branch header line: ## branch...tracking [ahead N, behind M]
        if lines and lines[0].startswith("## "):
            header = lines[0][3:]
            # Extract branch name (before ... or end)
            dot_idx = header.find("...")
            if dot_idx >= 0:
                branch = header[:dot_idx]
            else:
                # Could be "## branch" or "## No commits yet on branch"
                if header.startswith("No commits yet on "):
                    branch = header[len("No commits yet on "):]
                else:
                    branch = header.split()[0] if header else None
            # Parse ahead/behind
            bracket_idx = header.find("[")
            if bracket_idx >= 0:
                info = header[bracket_idx + 1:header.find("]")]
                for part in info.split(","):
                    part = part.strip()
                    if part.startswith("ahead "):
                        try:
                            ahead = int(part[6:])
                        except ValueError:
                            pass
                    elif part.startswith("behind "):
                        try:
                            behind = int(part[7:])
                        except ValueError:
                            pass

        # Parse file entries
        files: list[dict] = []
        for line in lines[1:]:
            if len(line) < 4:
                continue
            x = line[0]  # index status
            y = line[1]  # worktree status
            filepath = line[3:]

            # Rename entries quote each side separately ("old" -> "new"), so
            # split BEFORE unquoting would see the arrow inside quotes; the
            # porcelain arrow separator is never itself quoted, so splitting
            # first and unquoting each side is correct for both forms.

            # Handle renames/copies: "R  old -> new". Gate on the status
            # letters -- a plain modified file legitimately named
            # "foo -> bar" must NOT be split, or its row would point at an
            # unrelated file and clicking it edits the wrong one.
            if (x in ("R", "C") or y in ("R", "C")) and " -> " in filepath:
                filepath = filepath.split(" -> ", 1)[1]
            filepath = _porcelain_unquote(filepath)

            # Determine status code and staged flag
            if x == "?" and y == "?":
                files.append({"path": filepath, "status": "?", "staged": False})
            elif x == "!" and y == "!":
                continue  # ignored
            else:
                # If X is non-space/non-?, there's a staged change
                if x not in (" ", "?", "!"):
                    files.append({"path": filepath, "status": x, "staged": True})
                # If Y is non-space, there's an unstaged change
                if y not in (" ", "?", "!"):
                    files.append({"path": filepath, "status": y, "staged": False})

        # Merge numstat for line counts (staged + unstaged vs HEAD)
        try:
            numstat_rc, numstat_out, _ = _run_git_bounded(
                [*_git_cmd, "diff", "--numstat", "--no-textconv",
                 "--no-ext-diff", "HEAD"],
                cwd=base, env=_env, timeout=10,
            )
            if numstat_rc == 0:
                stats: dict[str, tuple[int | None, int | None]] = {}
                for ns_line in numstat_out.splitlines():
                    parts = ns_line.split("\t", 2)
                    if len(parts) == 3:
                        add_s, del_s, ns_path = parts
                        adds = int(add_s) if add_s != "-" else None
                        dels = int(del_s) if del_s != "-" else None
                        # numstat C-quotes the same class of paths status does;
                        # unquote so the merge key matches the parsed rows.
                        stats[_porcelain_unquote(ns_path)] = (adds, dels)
                for f in files:
                    if f["path"] in stats:
                        adds, dels = stats[f["path"]]
                        if adds is not None:
                            f["additions"] = adds
                        if dels is not None:
                            f["deletions"] = dels
        except FileNotFoundError:
            pass

        result: dict = {"repo": True, "repoRoot": repo_root, "files": files[:500]}
        if len(files) > 500:
            result["truncated"] = True
        if branch:
            result["branch"] = branch
        if ahead:
            result["ahead"] = ahead
        if behind:
            result["behind"] = behind
        return result

    result = await asyncio.to_thread(_run)
    # Egress redaction: repo content (paths, branch label, repo root) is
    # agent-influenceable and this response body is rendered by the dashboard,
    # so it goes through the same redaction as api_project_git. Normal values
    # pass through unchanged.
    if result.get("repoRoot"):
        result["repoRoot"] = redact(result["repoRoot"])
    if result.get("branch"):
        result["branch"] = redact(result["branch"])
    # Redact each file path, then drop entries that duplicate an earlier one
    # (preserving order and first occurrence). Same collision class as
    # api_project_tree: redact() can collapse two genuinely-different paths to
    # the same placeholder. This list feeds GitPanel, which keys its rows on
    # `${path}:${staged}` and takes its file total from files.length, so a
    # collision would render two indistinguishable rows under one React key and
    # overstate the count. (It cannot reach @pierre/trees as a duplicate the way
    # api_project_tree's list can: the tree's "changed" mode already collapses
    # status entries by path before handing them over.) The files[:500] cap was
    # already applied to the raw listing above, so this only removes collisions.
    #
    # The key is (path, status, staged), NOT path alone: one file with both
    # staged and unstaged changes ("MM", "AM", "MD") legitimately yields two
    # entries sharing a path but differing in status/staged, and GitPanel
    # renders them as separate rows. Keying on path alone would drop the
    # unstaged lane and undercount the file total. A real redaction collision
    # has an identical tuple, so it still collapses.
    deduped_files: list[dict] = []
    seen_keys: set[tuple[str, str | None, bool | None]] = set()
    for f in result.get("files", []):
        f["path"] = redact(f["path"])
        key = (f["path"], f.get("status"), f.get("staged"))
        if key in seen_keys:
            continue
        seen_keys.add(key)
        deduped_files.append(f)
    if "files" in result:
        result["files"] = deduped_files
    return web.json_response(result)


# Cap on entries returned by api_project_tree. The dashboard tree virtualizes
# rendering, so the cap bounds response size and walk time, not the UI.
_PROJECT_TREE_MAX_ENTRIES = 10_000

# Directories never worth listing in a workspace tree. Applied only on the
# non-git fallback walk — git listings already honor .gitignore.
_PROJECT_TREE_SKIP_DIRS = frozenset(
    {
        ".git",
        ".hg",
        ".svn",
        "node_modules",
        "__pycache__",
        ".venv",
        "venv",
        ".mypy_cache",
        ".pytest_cache",
        ".ruff_cache",
        ".tox",
        "dist",
        "build",
        ".next",
        ".cache",
        "target",
        ".gradle",
        ".idea",
    }
)


async def api_project_tree(request: web.Request) -> web.Response:
    """GET /api/project/tree?path=... - workspace file listing for a project dir.

    Returns project-relative POSIX file paths for rendering a workspace tree.
    Inside a git repository the listing is ``git ls-files --cached --others
    --exclude-standard`` scoped to the project dir (tracked + untracked,
    .gitignore honored); outside one it is a bounded directory walk. Path must
    match a known project directory (same allow-list as api_project_git).
    """
    state: DashboardState = request.app["state"]
    caller = request.get("user", "dashboard")
    raw = request.query.get("path", "").strip()
    if not raw:
        return web.json_response({"error": "path required", "code": "path_required"}, status=400)
    project = await asyncio.to_thread(
        _match_known_project_for, _slot_project_snapshot(state), raw
    )
    if project is None:
        _sel().log_api_access(
            caller=caller,
            operation="project_tree",
            outcome="denied",
            resources=raw,
            error="not a known project directory",
        )
        return web.json_response(
            {"error": "Unknown project directory", "code": "unknown_project_dir"}, status=403
        )

    base = await asyncio.to_thread(
        lambda: os.path.realpath(os.path.expanduser(project))
    )
    if await asyncio.to_thread(is_sensitive_path, base):
        _sel().log_api_access(
            caller=caller,
            operation="project_tree",
            outcome="denied",
            resources=base,
            error="sensitive path",
        )
        return web.json_response({"error": "Access denied", "code": "access_denied"}, status=403)
    _sel().log_api_access(
        caller=caller, operation="project_tree", outcome="allowed", resources=base
    )
    if not await asyncio.to_thread(os.path.isdir, base):
        return web.json_response({"root": redact(base), "paths": [], "repo": False})

    def _run() -> dict:
        # git listing first: honors .gitignore, includes tracked-but-deleted
        # files (they render with a deleted status lane), and with cwd=base a
        # project dir that is a repo SUBDIRECTORY lists only its own subtree.
        # -z: NUL separation, so no C-quoting and exotic names survive intact.
        probe_rc, _probe_out, _ = _run_git_bounded(
            ["git", "rev-parse", "--git-dir"], cwd=base, env=os.environ.copy(), timeout=5,
        )
        if probe_rc == 0:
            ls_rc, ls_out, _ = _run_git_bounded(
                # `core.fsmonitor=` disables the filesystem-monitor hook: it names a
                # command git would SPAWN, and it is repository-writable, so an agent
                # that can write `.git/config` could otherwise have a tree listing
                # execute it. Empty rather than `false` to match the sibling git
                # invocations in this module. The `rev-parse` probe above needs no
                # such guard — it reads no index and walks no working tree.
                [
                    "git", "-c", "core.fsmonitor=",
                    "ls-files", "-z", "--cached", "--others", "--exclude-standard",
                ],
                cwd=base,
                env=os.environ.copy(),
                timeout=15,
            )
            if ls_rc == 0:
                # SORT BEFORE THE CAP. `ls-files --cached --others` is not one
                # sorted stream: git emits every untracked entry as a complete
                # block and only then the tracked ones (its own emission order
                # -- unchanged if the flags are written the other way round, and
                # git-ls-files(1) documents no order at all). A prefix cut of
                # that therefore never reaches the tracked block once untracked
                # alone fill the cap, and the whole source tree loses its rows:
                # the dashboard infers a directory row only from the file paths
                # present, so those folders go absent rather than collapsed.
                # Sorting spends the budget by path instead of by whichever
                # block git happened to emit first. It does NOT make the two
                # branches emit the same order: the fallback walk below sorts
                # within each level but is depth-first overall, so it yields a
                # root `z.txt` before `a/x` where sorted() orders them the other
                # way. What the branches share is narrower and is the actual
                # warrant for sorting here -- this handler establishes its own
                # path order rather than passing through a source's arbitrary
                # emission order.
                listed = sorted(p for p in ls_out.split("\0") if p)
                truncated = len(listed) > _PROJECT_TREE_MAX_ENTRIES
                return {
                    "root": base,
                    "paths": listed[:_PROJECT_TREE_MAX_ENTRIES],
                    "repo": True,
                    "truncated": truncated,
                }

        # Fallback: bounded filesystem walk (non-repo project dirs).
        paths: list[str] = []
        truncated = False
        for dirpath, dirnames, filenames in os.walk(base):
            dirnames[:] = sorted(
                d for d in dirnames if d not in _PROJECT_TREE_SKIP_DIRS and not d.startswith(".")
            )
            rel_dir = os.path.relpath(dirpath, base)
            prefix = "" if rel_dir == "." else rel_dir.replace(os.sep, "/") + "/"
            for name in sorted(filenames):
                paths.append(prefix + name)
                if len(paths) >= _PROJECT_TREE_MAX_ENTRIES:
                    truncated = True
                    break
            if truncated:
                break
        return {"root": base, "paths": paths, "repo": False, "truncated": truncated}

    result = await asyncio.to_thread(_run)
    # Egress redaction, same rationale as api_project_git_status: listed names
    # are repo content and this body is rendered by the dashboard.
    result["root"] = redact(result["root"])
    # De-duplicate after redaction, preserving order and first occurrence.
    # redact() collapses each matched token to a fixed placeholder, so two
    # genuinely-different project-relative paths (e.g. a src/ vs target/ Maven
    # prefix and a credential-shaped filename token) can flatten to the same
    # redacted string. The dashboard tree hands this list straight to
    # @pierre/trees, whose appendPresortedPaths throws "Duplicate path" on
    # adjacent identical entries. dict.fromkeys keeps first occurrence. This
    # does not affect "truncated": the cap is applied to the raw listing above.
    result["paths"] = list(dict.fromkeys(redact(p) for p in result["paths"]))
    return web.json_response(result)


async def api_project_git_log(request: web.Request) -> web.Response:
    """GET /api/project/git/log?path=...&limit=N - recent commit log for a project dir.

    Returns short sha, subject, author, date (ISO), and isHead flag.
    Path must match a known project directory (same allow-list as api_project_git).
    """
    state: DashboardState = request.app["state"]
    caller = request.get("user", "dashboard")
    raw = request.query.get("path", "").strip()
    if not raw:
        return web.json_response({"error": "path required", "code": "path_required"}, status=400)

    limit_s = request.query.get("limit", "20")
    try:
        limit = max(1, min(100, int(limit_s)))
    except (ValueError, TypeError):
        limit = 20

    project = await asyncio.to_thread(
        _match_known_project_for, _slot_project_snapshot(state), raw
    )
    if project is None:
        _sel().log_api_access(
            caller=caller,
            operation="project_git_log",
            outcome="denied",
            resources=raw,
            error="not a known project directory",
        )
        return web.json_response({"error": "Unknown project directory", "code": "unknown_project_dir"}, status=403)

    base = await asyncio.to_thread(
        lambda: os.path.realpath(os.path.expanduser(project))
    )
    # Both probes stat the filesystem (a stalled network mount would block the
    # event loop), so they run in a worker thread like the realpath above.
    if await asyncio.to_thread(is_sensitive_path, base):
        _sel().log_api_access(
            caller=caller,
            operation="project_git_log",
            outcome="denied",
            resources=base,
            error="sensitive path",
        )
        return web.json_response({"error": "Access denied", "code": "access_denied"}, status=403)
    # Log the allow decision here (not after _run) so every authorized access
    # is audited, including the not-a-directory / not-a-repo early answers.
    _sel().log_api_access(
        caller=caller, operation="project_git_log", outcome="allowed", resources=base
    )
    if not await asyncio.to_thread(os.path.isdir, base):
        return web.json_response({"repo": False, "commits": []})

    def _run() -> dict:
        _git_cmd = [
            "git",
            "-c", "diff.textconv=",
            "-c", "core.attributesFile=/dev/null",
            "-c", f"core.hooksPath={os.devnull}",
            "-c", "core.fsmonitor=",
            # Repo-local .gitattributes is still consulted despite the
            # attributesFile override, so keep driver escape hatches shut and
            # emit non-ASCII paths raw (UTF-8) instead of C-quoted so the
            # panel can open them.
            "-c", "core.quotePath=false",
        ]
        _env = {**os.environ, "GIT_ATTR_NOSYSTEM": "1"}

        # Check if it's a repo
        probe_rc, _probe_out, _ = _run_git_bounded(
            [*_git_cmd, "rev-parse", "--git-dir"], cwd=base, env=_env, timeout=5,
        )
        if probe_rc != 0:
            return {"repo": False, "commits": []}

        # Same filter-driver refusal as the status handler (defense in depth:
        # ``git log`` does not run clean filters, but one uniform invariant --
        # no git subcommand runs against a repo that names a driver -- is
        # auditable; per-subcommand carve-outs are not).
        if _repo_declares_filter_driver(_git_cmd, base, _env):
            return {"repo": True, "commits": []}

        # Get HEAD sha for isHead marking
        head_rc, head_out, _ = _run_git_bounded(
            [*_git_cmd, "rev-parse", "--short", "HEAD"], cwd=base, env=_env, timeout=5,
        )
        head_sha = head_out.strip() if head_rc == 0 else ""

        # Separator unlikely in commit data
        sep = "\x1f"
        fmt = f"%h{sep}%s{sep}%an{sep}%aI"
        log_rc, log_out, _ = _run_git_bounded(
            [*_git_cmd, "log", f"--pretty=format:{fmt}", f"-{limit}"],
            cwd=base, env=_env, timeout=15,
        )
        if log_rc != 0:
            return {"repo": True, "commits": []}

        commits: list[dict] = []
        for line in log_out.splitlines():
            parts = line.split(sep, 3)
            if len(parts) < 4:
                continue
            sha, message, author, date = parts
            commits.append({
                "sha": sha,
                "message": message,
                "author": author,
                "date": date,
                "isHead": sha == head_sha,
            })
        return {"repo": True, "commits": commits}

    result = await asyncio.to_thread(_run)
    # Egress redaction: commit subjects and author names are repo content the
    # agent can author, and this body is rendered by the dashboard.
    for c in result.get("commits", []):
        c["message"] = redact(c["message"])
        c["author"] = redact(c["author"])
    return web.json_response(result)
